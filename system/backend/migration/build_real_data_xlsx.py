"""Отгоо эгчийн ГУРВАН ажлын дэвтрээс (.xlsx) `real_data.json` угсарна.

    python migration/build_real_data_xlsx.py --src "…/Түрээсийн тооцоо 2026 2"

Эх сурвалж (P0-10):
  WB1 «Түрээсийн тооцоо 2026.xlsx»        — хөдөлгүүр: харилцагч тус бүрийн
        хуудас, хэвтээ циклийн блокууд (тоо·үнэ·хоног·дүн), тооллого 6.22
  WB2 «Үндсэн материалын тооцоо 24он.xlsx» — паркийн дэвтэр: МӨР=харилцагч ×
        БАГАНА=SKU гадаа матриц («2026 шинэ» = хамгийн сүүлийн байдал)
  WB3 «2026 тооцоо.xlsx»                   — удирдлагын самбар: «Түрээс
        тооцоо-26» (Нийт дүн · Тооцоо хийсэн · **Үлдэгдэл** · Барьцаа),
        «Өглөгө.зээл-26» (зээл)

ЗАРЧИМ (docs/Чадварын харьцуулалт.md §1):
  * Самбарын `Үлдэгдэл` бол ТҮҮНИЙ эрх мэдэлтэй авлага — бид түүнийг л ачаална.
  * `Барьцаа` балансын ГАДНА (R21); «байршуулаагүй» = 0 биш, БАЙХГҮЙ.
  * Сөрөг үлдэгдэл = «илүү» (R18) → кредит төлбөр, өр биш.
  * Нэг баримт гурван дэвтэрт ЗӨРӨХ эрхтэй (олдвор б) — бид зөрүүг НУУХГҮЙ,
    тайланд гаргаж, шийдвэрийг ЭЗЭНД нь үлдээнэ.
  * Эргэлзээтэй нэрийг НЭГТГЭХГҮЙ — тусад нь үлдээж туг өргөнө.
  * Тааруулж чадаагүй юмыг чимээгүй хаяхгүй — warnings-д гарна.

Гаралт нь `app/services/migration.py:load_data()`-ийн уншдаг бүтэц + `audit`
блок (тулгалтын тайлан үүнээс уншина). Детерминистик: ижил оролт → ижил байт.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime

try:
    from barter_manual import BARTER
except ImportError:  # pragma: no cover — багцаас дуудахад
    from .barter_manual import BARTER  # type: ignore

# Дэвтрүүдийн байдлаар: хамгийн сүүлийн бичилт 2026.08.04 → 9-р сарын 1-нээс
# систем тоолж эхэлнэ. Тогтмол — «өнөөдөр» гэвэл гаралт өдөр бүр өөрчлөгдөнө.
DEFAULT_AS_OF = "2026-09-01"

WB1 = "Түрээсийн тооцоо 2026.xlsx"
WB2 = "Үндсэн материалын тооцоо 24он.xlsx"
WB3 = "2026 тооцоо.xlsx"


# ════════════════════════════════════════════════════════ 1. ЦЭВЭР ФУНКЦҮҮД
# (Excel-гүйгээр тесттэй — tests/test_build_real_data.py)

_SCALE = {"к": 1e3, "мянга": 1e3, "сая": 1e6, "тэрбум": 1e9}
_DATEISH = re.compile(r"^\s*\d{4}\s*[./-]\s*\d{1,2}\s*(?:[./-]\s*\d{1,4})?\s*$")
_NUM = re.compile(r"(?<!\w)(-?\d[\d   ,]*(?:\.\d+)?)")
_SCALE_RE = re.compile(r"\s*(к|К|мянга|сая|тэрбум)", re.IGNORECASE)
_WORD_START = re.compile(r"\s*[^\W\d_]", re.UNICODE)


def parse_money(v) -> float | None:
    """Нүднээс ₮. Товчлол «16.5к»=16,500 · «170сая»=170,000,000.

    «6м» бол ТРУБАНЫ УРТ — мөнгө биш. Тоо дараа нь танигдаагүй үг ирвэл
    эсвэл ард нь дахин тоо (циклийн шошго «4.14-5.13») ирвэл None.
    """
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (datetime, date)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(" ", " ").replace(" ", " ").strip()
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1].strip()
    s = re.sub(r"₮|төгрөг|төг\b", " ", s, flags=re.IGNORECASE).strip()
    if _DATEISH.match(s):
        return None
    m = _NUM.search(s)
    if not m:
        return None
    rest = s[m.end():]
    mult = 1.0
    ms = _SCALE_RE.match(rest)
    if ms:
        mult = _SCALE[ms.group(1).lower()]
        rest = rest[ms.end():]
    elif _WORD_START.match(rest):
        return None                      # тоо + үг = хэмжээс/код, мөнгө биш
    if re.search(r"\d", rest):
        return None                      # «4.14-5.13», «№25/04»
    try:
        val = float(m.group(1).replace(" ", "").replace(",", ""))
    except ValueError:
        return None
    val *= mult
    return -val if neg else val


_D_YMD = re.compile(r"^(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})(?!\d)")
_D_DMY = re.compile(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})\s*$")
_D_YM = re.compile(r"^(\d{4})[.\-/](\d{1,2})\s*$")
_D_MD = re.compile(r"^(\d{1,2})[.](\d{1,2})\s*$")


def parse_date_any(v, default_year: int | None = None) -> date | None:
    """Түүний ТАВАН огнооны хэлбэр:
    2026.07.20 · 2026.1.22 · 2026-06-25(ISO/datetime) · 2026/4(сар) · 25.06.2024
    Зургаа дахь нь «7.05» (сар.өдөр) — жилгүй тул `default_year` шаардана.
    """
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None

    def _mk(y, m, d):
        try:
            return date(int(y), int(m), int(d))
        except ValueError:
            return None

    m = _D_YMD.match(s)
    if m:
        return _mk(m.group(1), m.group(2), m.group(3))
    m = _D_DMY.match(s)
    if m:
        return _mk(m.group(3), m.group(2), m.group(1))
    m = _D_YM.match(s)
    if m:
        return _mk(m.group(1), m.group(2), 1)
    m = _D_MD.match(s)
    if m and default_year:
        return _mk(default_year, m.group(1), m.group(2))
    return None


_LEGAL = re.compile(r"(ххк|ххн|ххт|хк|llc|ltd)")


def name_key(s: str) -> str:
    """Нэрийн харьцуулах түлхүүр — том/жижиг, зай, хуулийн хэлбэр, хашилт,
    э/е ба ё/е-гийн хэлбэлзлийг арилгана («Повер»↔«Повэр»)."""
    s = unicodedata.normalize("NFKC", str(s or "")).casefold()
    s = re.sub(r"[«»\"'`.,\-–—()]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = _LEGAL.sub(" ", s)
    s = s.replace(" ", "")
    return s.replace("э", "е").replace("ё", "е")


# ── Нэрийн хувилбарууд → ТҮҮНИЙ самбарын нэр ────────────────────────────────
# Түлхүүр = каноник (AR самбар «Түрээс тооцоо-26»), утга = баримтжсан хувилбар.
# Нэмэх бүрд ЯАГААД гэдгийг бич — энэ хүснэгт нь тайланд хэвлэгддэг.
ALIASES: dict[str, list[str]] = {
    "Өнө Орд ХХК яармаг": ["ӨнөОрд", "ӨнөОрд ХХК", "Өнө Орд ХХК", "өнөорд",
                           "ӨнөОрд ХХК архив", "өнөорд-Архив", "ӨнөОрд-8"],
    "Өнө Орд ХХК нүхт": ["Өнөорд Нүхт", "ӨнөОрд ХХК нүхт", "Өнө Орд ХХК Нүхт",
                         "Өнөорд Нүхт-20"],
    "Марч констракшн": ["Марч", "Марч констракшн ХХК", "Марч-1"],
    "Блүүм технологи": ["БЛҮҮМ технологи", "БЛҮҮМ технологи ХХК", "БЛҮҮМ ХХК",
                        "БЛҮҮМ-2", "Блүүт тооцоо"],
    "Грэйт Майнинг": ["Грэйт майнинг", "Грэйт Майнинг ХХК", "ГрэйтМайнинг ХХК",
                      "ГрэйтМайнинг-5"],
    "Бутангууд": ["Бутангууд ХХК", "Бутангууд констракшн ХХК", "бутан",
                  "Бутангууд-7"],
    # хуудасны толгой: «"Гоби Ресурс Хилл" ХХК Хурд зайсан төсөл» = самбарын
    # «Говь Ресурс Хурд» (нэг төсөл, нэг гэрээ №25/21)
    "Говь Ресурс Хурд": ["Гоби Ресурс Хилл", "Гоби Ресурс Хилл ХХК", "Говьресурс",
                         "ГобиресурсХилл", "Говьресурс-4", "говь тооцоо"],
    "Монуул ХХК": ["Монуул", "Монуул МЧМ майнинг ХХК", "Монуул-10", "Монуул тооцоо"],
    "БМонгол ХХК": ["БМонгол", "Билгүүн Монгол ХХК", "БМ-12"],
    "Зулаа": ["иргэн О.Зулаа", "О.Зулаа", "Зулаа-3"],
    "Мөнхболд": ["Мөнхболд-9"],
    "Ашид Донж Билгүүн": ["Ашид донж билгүүн", "Ашид Донж Билгүүн ХХК", "АшидДонж-11"],
    "Удам констракшн": ["Удам констракшн ХХК", "Удам-13"],
    "Шижир Тиара": ["Шижир Тиара ХХК", "ШижирТиара-14"],
    # Самбарт «Эдэнэс» гэж бичигдсэн (үсэг унасан) — самбар нь эх сурвалж тул
    # каноник нь ТҮҮНИЙ бичсэнээр үлдэнэ.
    "Агар Эдэнэс Трэйд": ["Агар Эрдэнэс Трэйд", "Агар Эрдэнэс Трэйд ХХК",
                          "Агар Эрдэнэс-15"],
    "Арвинбулаг констракшн": ["Арвинбулаг", "Арвинбулаг констракшн ХХК",
                              "Арвинбулаг-16"],
    "Сүхжин ХХК": ["Сүхжин", "Сүхжин-18"],
    "Хатанбүүвэй": ["Д.Хатанбүүвэй", "Хатанбүүвэй-19"],
    "Цэрэнчимэд": ["Н.Цэрэнчимэд", "Цэрэнчимэд-21"],
    "Нью Мэн Повер ХХК": ["Нью Мэн Повер", "НьюМэнПовэр", "НьюМэнПовэр-22"],
    "Өрнүүн Эрчит Зам ххк": ["Өрнүүн Эрчит Зам", "Өрнүүн эрчит зам ХХК",
                             "Өрнүүн эрчит зам-17"],
    "Орхонзам": ["ОрхонЗам ХХК"],
    "Голден лайт": ["Голден лайт групп"],
    "Тэнгис тэмүүлэгч": ["Тэнгис Тэмүүлэгч ХХК", "ТэнгисТэмүүлэгч"],
    "Түдэвтэй Уул": ["Түдэвтэй уул"],
    "Жаргалант Ургамал": ["Жаргалант Ургамал ХХК", "Жаргалант ургамал-23"],
    "Бэлгүтэй": ["Х.Бэлгүтэй", "Бэлгүтэй-6"],
    "Эрдэнэт шанд": ["иргэн Ч.Баярмагнай", "Ч.Баярмагнай", "эрдэнэт шанд-24"],
    # ── зөвхөн AR самбарт: хувилбаргүй, гэхдээ каноник ──
    "Дархан Оюунаа": [], "Тасралтгүй урсгал": [], "Грандслаб": [],
    "Хурд групп": [],
    # ── зөвхөн ПАРКИЙН дэвтэрт (WB2): бие даасан тал гэж үзэв, нэгтгээгүй ──
    "БЛҮҮМ архангай": [], "Блүүм дарь эх": [], "Говьресурс УТБА": [],
    "Говьресус БМ": [], "Хурд Гранд Слаб": [], "Хурд Бетон арматур": [],
    "УранОчир": [], "Батхолбоо": [], "Сачи-бэхи 38": [], "Ганцоож": [],
    "Угаал Гол ХХК": [], "Орон нутаг": [], "Ангирмаа": [], "Д.Мягмаржав": [],
    "Флорес импекс ХХК": [], "Зүмбэр харш": [], "Оюунаа эгч": [],
}

# Нэгтгээгүй — ТҮҮНИЙ шийдэх ёстой хосууд. (нэр А, нэр Б, асуулт)
AMBIGUOUS: list[tuple[str, str, str]] = [
    ("Блүүм технологи", "БЛҮҮМ архангай",
     "Архангайн талбай нь Блүүмийн НЭГ авлагад орох уу, тусдаа харилцагч уу? "
     "НОТОЛГОО: БЛҮҮМ-2 хуудасны 4,294ш = технологи 2,044 + архангай 326 + "
     "дарь эх 1,924 — яг таарч байна, тэгэхээр хуудас нь ГУРВУУЛАНГ хамарсан."),
    ("Блүүм технологи", "Блүүм дарь эх",
     "Дарь эхийн талбай нь Блүүмийн НЭГ авлагад орох уу, тусдаа харилцагч уу? "
     "(дээрх нотолгоог үзнэ үү)"),
    ("Говь Ресурс Хурд", "Говьресурс УТБА",
     "УТБА-гийн төсөл нь ижил тал уу? (паркийн дэвтэрт тусдаа мөр)"),
    ("Говь Ресурс Хурд", "Говьресус БМ",
     "«Говьресус БМ» нь Говьресурсын өөр төсөл үү, БМонгол уу?"),
    ("Хурд групп", "Грандслаб",
     "«Хурд Гранд Слаб» нь Хурд группын дотор уу, тусдаа «Грандслаб» уу?"),
    ("Дархан Оюунаа", "Оюунаа эгч",
     "Паркийн дэвтрийн «Оюунаа эгч» нь самбарын «Дархан Оюунаа» мөн үү?"),
    ("Бутангууд", "Батцоож",
     "«Батцоож» хуудас нь Бутангуудын дэд дэвтэр үү (гэрээ №25/19 ижил)?"),
]

_ALIAS_INDEX: dict[str, str] = {}
for _canon, _variants in ALIASES.items():
    _ALIAS_INDEX[name_key(_canon)] = _canon
    for _v in _variants:
        _ALIAS_INDEX[name_key(_v)] = _canon


@dataclass(frozen=True)
class Canon:
    name: str
    matched: bool
    raw: str


def canon_client(raw) -> Canon:
    """Нэрийн хувилбар → каноник нэр. Танигдаагүйг ХЭВЭЭР нь буцаана
    (`matched=False`) — чимээгүй хаях нь мөнгө алдах хамгийн хямд арга."""
    s = re.sub(r"\s+", " ", str(raw or "")).strip()
    key = name_key(s)
    if key in _ALIAS_INDEX:
        return Canon(_ALIAS_INDEX[key], True, s)
    return Canon(s, False, s)


def parse_deposit(v) -> float | None:
    """Барьцаа. «байршуулаагүй» → None (0 гэвэл ТӨЛСӨН мэт уншигдана, R21)."""
    if isinstance(v, str) and "байршуул" in v.casefold():
        return None
    return parse_money(v)


@dataclass(frozen=True)
class Balance:
    debt: float
    credit: float


def split_balance(x: float | None) -> Balance:
    """Сөрөг үлдэгдэл = ИЛҮҮ төлөлт (R18) — өр биш, кредит."""
    v = float(x or 0)
    return Balance(debt=v if v > 0 else 0.0, credit=-v if v < 0 else 0.0)


def reconcile_balance_and_deposit(total, paid, balance, deposit):
    """Самбарын мөрөөс (үлдэгдэл, барьцаа, тэмдэглэл).

    БМонголын кейс: `Үлдэгдэл` хоосон, сөрөг дүн нь БАРЬЦААНЫ баганад
    бичигдсэн — барьцаа гэж уншвал 192,500₮ хоёр газраас нэг зэрэг гарна.
    """
    note = ""
    bal = balance
    if bal is None:
        bal = (total or 0) - (paid or 0)
    dep = deposit or 0
    if bal < 0 and dep > 0 and abs(abs(bal) - dep) < 1:
        dep = 0
        note = ("Сөрөг үлдэгдэл БАРЬЦААНЫ баганад бичигдсэн — барьцаа гэж "
                "тооцоогүй, илүү төлөлт (кредит) болгов")
    return float(bal), float(dep), note


# ── Материалын код ──────────────────────────────────────────────────────────
# Каталогт УРЬД нь байсан кодууд (seed.py).
HEV_CODES = ("6012", "5012", "4512", "4012", "3012", "2012",
             "1010", "1015", "1020", "1515", "1200", "2400")
TRUBA_CODES = ("6м", "4м", "3м", "2м")
# ⚠ Дэвтэрт бий — каталогт АЛГА байсан (E3). ЧИМЭЭГҮЙ УНАХАА БОЛИВ: эдгээр нь
# одоо каталогт НЭЭГДЭНЭ (`catalog` блок). Өнө Ордын «1м» 278ш яг эндээс унадаг
# байв.
HEV_GAP_CODES = ("1025", "2020")
TRUBA_GAP_CODES = ("5м", "1.5м", "1м")
GRADE_SUFFIX = {"пл": "плас", "плас": "плас", "пластик": "плас",
                "ш": "шинэ", "шинэ": "шинэ",
                "а": "А", "в": "В", "b": "В", "сольсон": "сольсон", "шорт": "шорт"}

#: Каталогт НЭЭГДЭХ шинэ материалууд (§4 E3). `base_rate` нь ТҮҮНИЙ тарифын
#: хүснэгтээс (Чадварын харьцуулалт §1 R2 «труба уртаараа», булан = хэвийн
#: тариф 330₮) гарган авсан; гаргаж чадаагүйг 0 + ТУГТАЙ тэмдэглэл.
#:
#: ⚠ `repair_fee` ХААНА Ч БАЙХГҮЙ. Гурван дэвтэрт «засвар» гэсэн үг НИЙТ НЭГ
#: удаа, чөлөөт тэмдэглэл болж гарна — засварын хураамжийн ХҮСНЭГТ түүнд алга.
#: Урьд нь энд 4,000₮ ба 8,000₮ бичигдсэн байсан нь ЗОХИОСОН тоо байв: ачаагч
#: бүх материалын `repair_fee`-г 0 болгодог тул эндээс ч 0-оор л гарна.
CATALOG_NEW: dict[str, dict] = {
    "Хэв хашмал 1025": {"category": "Хэв", "base_rate": 330, "repair_fee": 0,
                        "why": "Дотор булан 1025 — булангийн тариф 330₮ (R2)"},
    "Хэв хашмал 2020": {"category": "Хэв", "base_rate": 330, "repair_fee": 0,
                        "why": "Дотор булан 2020 — булангийн тариф 330₮ (R2)"},
    "Труба 1м": {"category": "Труба", "base_rate": 110, "repair_fee": 0,
                 "why": "ӨнөОрд-8!AQ27 = 110₮ (түүний өөрийн тариф)"},
    "Труба 5м": {"category": "Труба", "base_rate": 0, "repair_fee": 0,
                 "why": "Тариф ХААНААС Ч олдсонгүй — 0-оор нээв, ТЭР тогтооно"},
    "Труба 1.5м": {"category": "Труба", "base_rate": 0, "repair_fee": 0,
                   "why": "Тариф ХААНААС Ч олдсонгүй — 0-оор нээв, ТЭР тогтооно"},
    "Шат": {"category": "Бусад", "base_rate": 0, "repair_fee": 0,
            "why": "Ангилал нь ч нээгдээгүй — 0-оор нээв, ТЭР тогтооно"},
}

# Дэвтэрт бий — каталогт АЛГА байсан кодын тайлбар (тайланд хэвлэгдэнэ).
CATALOG_GAPS: dict[str, str] = {
    "1025": "Дотор булан 1025 — каталогт нээв (Хэв хашмал 1025)",
    "2020": "Дотор булан 2020 — каталогт нээв (Хэв хашмал 2020)",
    "5м": "Труба 5м — каталогт нээв (тариф 0)",
    "1.5м": "Труба 1.5м — каталогт нээв (тариф 0)",
    "1.5": "Труба 1.5м — каталогт нээв (тариф 0)",
    "1м": "Труба 1м — каталогт нээв (тариф 110₮)",
    "Шат": "Шат — каталогт нээв (тариф 0, ангилал «Бусад»)",
}


def grade_token(raw) -> str | None:
    """Нүдэнд БҮХЭЛДЭЭ зэрэглэлийн үг байвал зэрэглэлийг буцаана (E3).

    Тэр зэрэглэлээ КОДЫН биш НЭРИЙН баганад бичдэг: `ӨнөОрд-8!AN8='пластик'`
    (`AO8=6012`), `Бутангууд-7!T26='шинэ'` (`U26=1010`). Эдгээр 3 үг өмнө нь
    `А` болж нурж, 321 + 158 + 31 + 36ш буруу зэрэглэлээр ачаалагдаж байв.
    """
    if raw is None or isinstance(raw, bool) or isinstance(raw, (int, float)):
        return None
    return GRADE_SUFFIX.get(str(raw).strip().casefold())


def material_of(raw) -> tuple[str, str] | None:
    """Нүд → (каталогийн нэр, зэрэглэл) эсвэл None.
    6012→Хэв · В2/V2→Тулаас · 6м→Труба. «5012 пл», «V2 шинэ» → зэрэглэл."""
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, float) and raw == int(raw):
        s = str(int(raw))
    else:
        s = str(raw).strip()
    if not s:
        return None
    s = re.sub(r"(\d)\s+м\b", r"\1м", s)          # «4 м» → «4м»
    parts = s.split()
    base = parts[0]
    base = re.sub(r"^[BbVvВв](?=\d)", "В", base)
    grade = GRADE_SUFFIX.get(parts[1].casefold(), "А") if len(parts) > 1 else "А"
    if base in HEV_CODES or base in HEV_GAP_CODES:
        return (f"Хэв хашмал {base}", grade)
    if re.fullmatch(r"В[2-6]", base):
        return (f"Тулаас {base}", grade)
    if re.fullmatch(r"\d+(?:\.\d+)?м", base):
        return ((f"Труба {base}", grade)
                if base in TRUBA_CODES or base in TRUBA_GAP_CODES else None)
    if base.casefold() == "шат":
        return ("Шат", grade)
    return None


# ── Гэрээний толгой, цикл, НӨАТ, барьцаа, гарын үсэг (цэвэр функцүүд) ───────

_HEAD_DATE = re.compile(r"(\d{4})\s*[.\-/]\s*(\d{1,2})\s*[.\-/]\s*(\d{1,2})")
_APPENDIX = re.compile(r"хавсралт\s*№?\s*(\d{1,2})", re.IGNORECASE)


def parse_contract_header(text) -> dict:
    """`'2024.4.04 №24/ 03  Гэрээний хавсралт № 4'` → огноо · № · хавсралт №.

    Гэрээ бүр ЖИНХЭНЭ эхлэлтэй: Марч 2022.3.1, Блүүм 2024.4.04, Грэйт
    2025.5.08, Бутангууд/ӨнөОрд 2025.09.22, Ашид 2026.03.25. Урьд нь БҮГД
    `as_of` (2026-09-01) болж, дөрвөн жилийн харилцаа өнөөдрөөс эхэлдэг байв.
    """
    s = re.sub(r"\s+", " ", str(text or "")).strip()
    out = {"date": None, "no": "", "appendix": "", "raw": s}
    if not s:
        return out
    m = _HEAD_DATE.search(s)
    if m:
        try:
            out["date"] = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            out["date"] = None
    # № нь огнооны ДАРАА ирнэ; «№ 4» (хавсралт) биш «№24/03» (гэрээ)
    m = _NO_RE.search(s) or _NO_FALLBACK.search(s)
    if m:
        out["no"] = re.sub(r"\s+", "", m.group(1))
    m = _APPENDIX.search(s)
    if m:
        out["appendix"] = m.group(1)
        if out["no"] == out["appendix"]:
            out["no"] = ""                       # «Гэрээний хавсралт №4»-ийг л барив
    return out


_MONTH_LEN = (31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31)


def cycle_mode_from_labels(labels) -> str:
    """Циклийн шошгонууд → `"month"` эсвэл `"days"` (R5 / H3).

    `'4.01-4.30'`, `'5.01-5.31      31х'`, `'6.01-6.30'` — эхлэл нь сарын 1,
    төгсгөл нь тэр сарын СҮҮЛЧИЙН өдөр → КАЛЕНДАРЬ САР. `'3.15-4.13'` бол
    30 хоногийн зангилаа. Грэйт Майнингийн сар бүр 1,363,320₮ дутуу
    нэхэгдэж байсан нь ЭНЭ ялгаанаас.
    """
    good = bad = 0
    for lab in labels or []:
        m = _CYCLE_RE.search(_txt(lab))
        if not m:
            continue
        m1, d1, m2, d2 = (int(m.group(i)) for i in (1, 2, 3, 4))
        if not (1 <= m1 <= 12 and 1 <= m2 <= 12):
            continue
        if d1 == 1 and m1 == m2 and d2 == _MONTH_LEN[m1 - 1]:
            good += 1
        else:
            bad += 1
    return "month" if good and good > bad else "days"


def last_covered_from_labels(labels, year: int) -> date | None:
    """Дэвтэр ХААНА ХҮРТЭЛ нэхэгдсэн бэ — СҮҮЛЧИЙН циклийн мөрийн ТӨГСГӨЛ.

    `'7.13-8.11'` → 2026-08-11 (Блүүм); календарь горимд `'8.01-8.31'` →
    8-р сарын сүүлчийн өдөр. Энэ огноо нь `billing_from`-ийн эх сурвалж:
    систем ЯГ маргааш нь тоолж эхэлнэ.

    ХАМГИЙН ИХИЙГ нь БИШ, СҮҮЛЧИЙНХИЙГ нь авна — Ашид Донжийн хуудсанд
    `'7.23-8.21'`-ийн ДАРАА `'8.01-8.31'` бичигдсэн; дараалал нь ТҮҮНИЙХ.

    Жил нь шошгонд бичигдээгүй тул дэвтрийн жилээр (`year`) тавина; төгсгөл
    нь эхлэлээсээ ӨМНӨХ сард унавал (`'12.15-1.13'`) дараагийн жил болно.
    Уншигдахгүй бол ТААМАГЛАХГҮЙ — None (ачаалагч тугтай мөр өргөнө).
    """
    last = None
    for lab in labels or []:
        m = _CYCLE_RE.search(_txt(lab))
        if not m:
            continue
        m1, d1, m2, d2 = (int(m.group(i)) for i in (1, 2, 3, 4))
        if not (1 <= m1 <= 12 and 1 <= m2 <= 12):
            continue
        try:
            last = date(year + (1 if m2 < m1 else 0), m2, d2)
        except ValueError:
            continue                             # «2.30» гэх мэт байхгүй өдөр
    return last


_BOARD_MONTH = re.compile(r"^(\d{1,2})\s*сар$", re.IGNORECASE)


def board_month_columns(header) -> dict[int, int]:
    """Самбарын толгойн `'3 сар' … '7 сар'` → {баганын индекс: сарын дугаар}."""
    out: dict[int, int] = {}
    for ci, c in enumerate(tuple(header or ())):
        m = _BOARD_MONTH.match(_txt(c))
        if m and 1 <= int(m.group(1)) <= 12:
            out[ci] = int(m.group(1))
    return out


def board_last_covered(header, row, year: int) -> date | None:
    """САМБАР хаана хүртэл нэхсэн бэ — дүн бичигдсэн СҮҮЛЧИЙН сарын багана.

    Самбар нь сарын нэхэлтийг нэг нүдэнд хийдэг тул хамралтыг ОЙРОЛЦООГООР
    (сарын сүүлчээр) л хэлнэ — харилцагчийн хуудасны `last_covered` нь
    нарийвчлалтай. Хоёр нь зөрвөл систем аль нэгийг СОНГОХГҮЙ: тайлангийн
    §9-д хоёуланг нь, зөрүүг нь хоног ба ₮-өөр бичээд ТЭР шийднэ.
    """
    cols = board_month_columns(header)
    if not cols:
        return None
    cells = tuple(row or ()) + (None,) * (max(cols) + 1)
    best = None
    for ci, mon in sorted(cols.items()):
        if parse_money(cells[ci]):
            best = mon
    if best is None:
        return None
    return date(year, best, monthrange(year, best)[1])


VAT_PERCENT_RE = re.compile(r"нөат\s*(\d{1,2})\s*%", re.IGNORECASE)


def vat_of(texts) -> tuple[float, list[str]]:
    """Хуудсуудын БҮХ текстээс НӨАТ-ын хувь + тэмдэглэл (R14 / H12).

    «НӨАТ 10%» гэсэн МӨР бол хувь нь 10; «нөат шивсэн», «НӨАТ харилцан
    шивэлцэнэ», «нөат авах бол 10% нэмж төлбөр хийнэ» бол ФАКТ — гэрээн дээр
    тэмдэглэл болно, хувь хэмжээг ТЭР тогтооно.
    """
    pct, notes = 0.0, []
    for t in texts or []:
        s = _txt(t)
        if "нөат" not in s.casefold():
            continue
        m = VAT_PERCENT_RE.search(s)
        if m and re.fullmatch(r"\s*нөат\s*\d{1,2}\s*%\s*", s, re.IGNORECASE):
            pct = max(pct, float(m.group(1)))
        elif s not in notes:
            notes.append(s)
    return pct, notes


_REF_CHAIN = re.compile(r"^=\+?[A-Z]{1,3}\d{1,5}(?:\s*\+\s*[A-Z]{1,3}\d{1,5})*$")
_REF = re.compile(r"([A-Z]{1,3})(\d{1,5})")
_SUM_RANGE = re.compile(r"^=SUM\(([A-Z]{1,3})(\d{1,5}):([A-Z]{1,3})(\d{1,5})\)$",
                        re.IGNORECASE)


def parse_total_formula(formula) -> dict | None:
    """ТҮҮНИЙ «Нийт» тооны нүд нь АЛЬ МӨРҮҮДИЙГ тоолж байгааг хэлнэ (E1).

    Хоёр хэлбэр:
      `=SUM(V7:V60)`                    → муж (тэр блокийн бүх мөр)
      `=+V58+V55+V52+…+V10`             → ГАРААР ТҮҮСЭН 16 мөр (Бутангууд V61)

    Хоёр дахь нь ШИЙДВЭР: буцаагдсан дэд мөрүүдийг тэр ӨӨРӨӨ хассан. Урьд нь
    задлагч түүнийг үл ойшоож, огноогүй буцаалт бүрийг ГАДАА гэж тоолдог байв
    — Бутангууд 4,432ш болж хэтэрсэн (түүний өөрийн тоо 1,879).
    """
    s = re.sub(r"\s+", "", str(formula or ""))
    if not s.startswith("="):
        return None
    m = _SUM_RANGE.match(s)
    if m:
        lo, hi = int(m.group(2)), int(m.group(4))
        return {"kind": "range", "rows": list(range(min(lo, hi), max(lo, hi) + 1))}
    if _REF_CHAIN.match(s):
        return {"kind": "refs", "rows": [int(r) for _, r in _REF.findall(s)]}
    return None


def sign_from_formula(formula, default: int = 1) -> int:
    """Үлдэгдлийн томьёоны ТЭМДЭГ (§4 / №106).

    Самбарын БҮХ мөр `«=+H8-I8»` (нэхэмжилсэн − төлсөн). Хурд группын
    `J32 «=+I32-H32»` ГАНЦААРАА ЭСРЭГ бичигдсэн — 78,165,000₮ нь ӨР биш,
    ИЛҮҮ ТӨЛӨЛТ. Задлагч тэмдгийг залгидаг тул 156.3 сая₮-ийн эргэлт гардаг.
    """
    s = re.sub(r"\s+", "", str(formula or "")).upper()
    m = re.match(r"^=\+?([A-Z]{1,3})(\d+)-([A-Z]{1,3})(\d+)$", s)
    if not m:
        return default
    a, b = m.group(1), m.group(3)
    if a == "I" and b == "H":
        return -1
    if a == "H" and b == "I":
        return 1
    return default


_CHAIN_TERM = re.compile(r"([+-]?)\s*(\d[\d.]*)")


def parse_money_chain(formula) -> list[float] | None:
    """`«=20000000-8265000+3000000+3000000+10000000»` → 5 гишүүн (R22 / H8).

    Гишүүн бүр нь ШИЙДВЭР: байршуулав · суутгав · нэмж байршуулав. Нэг float
    болж нурахад «аль нь суутгагдсан» гэдэг алга болно.
    """
    s = re.sub(r"\s+", "", str(formula or ""))
    if not s.startswith("=") or not re.fullmatch(r"=[+-]?[\d.+\-]+", s):
        return None
    body = s[1:]
    out: list[float] = []
    for sign, num in _CHAIN_TERM.findall(body):
        try:
            v = float(num)
        except ValueError:
            return None
        out.append(-v if sign == "-" else v)
    return out or None


def deposit_events_from_chain(terms, as_of, ref: str = "") -> list[dict]:
    """Барьцааны гинж → дэвтрийн ЯВДЛУУД (lodge/topup/apply).

    ЭЕРЭГ гишүүн: эхнийх нь `lodge`, дараагийнх нь `topup`.
    СӨРӨГ гишүүн: `apply` — авлагад суутгасан. ⚠ Шинэ ТӨЛБӨР ҮҮСГЭХГҮЙ:
    самбарын `Үлдэгдэл` тэр суутгалыг АЛЬ ХЭДИЙН цэвэрлэсэн (давхар тооцвол
    авлага хоёр дахин буурна).
    """
    out: list[dict] = []
    lodged = False
    for t in terms or []:
        if not t:
            continue
        if t > 0:
            out.append({"kind": "lodge" if not lodged else "topup", "amount": t,
                        "date": str(as_of), "note": f"хуучин системээс — {ref}"})
            lodged = True
        else:
            out.append({"kind": "apply", "amount": -t, "date": str(as_of),
                        "note": f"хуучин системээс — {ref}"})
    return out


_PHONE = re.compile(r"(?<!\d)(\d{8})(?!\d)")
ROLE_WORDS = ("гүйцэтгэх захирал", "төслийн менежер", "талбайн менежер",
              "захирал", "нярав", "менежер", "түрээслэгч", "түрээлэгч")


def parse_signatory(text, phone_cells=()) -> dict | None:
    """Гарын үсгийн мөр → {нэр, албан тушаал, утас, утас2} (№72, 73).

    `'Төслийн менежер: Н.Батцоож ............................'` + 96590908
    `' Захирал Б.Дарханбаяр  ........................  88111935  99991491'`
    `'иргэн О.Зулаа ...................   99205311 '`
    """
    s = re.sub(r"\s+", " ", str(text or "")).strip()
    if not s:
        return None
    low = s.casefold()
    role = ""
    for w in ROLE_WORDS:
        if w in low:
            role = w.capitalize()
            break
    phones = _PHONE.findall(s)
    for c in phone_cells:
        if c is None:
            continue
        phones += _PHONE.findall(str(int(c)) if isinstance(c, (int, float)) else str(c))
    name = s
    if ":" in name:
        name = name.split(":", 1)[1]
    elif role and low.startswith(role.casefold()):
        name = name[len(role):]
    name = re.sub(r"[.…]{2,}", " ", name)
    name = _PHONE.sub(" ", name)
    name = re.sub(r"\s+", " ", name).strip(" .,:;-")
    if role and name.casefold().startswith(role.casefold()):
        name = name[len(role):].strip(" .,:;-")
    if not name or len(name) > 60:
        return None
    if not re.search(r"[^\W\d_]", name, re.UNICODE):
        return None
    seen: list[str] = []
    for p in phones:
        if p not in seen:
            seen.append(p)
    return {"name": name, "role": role,
            "phone": seen[0] if seen else "", "phone2": seen[1] if len(seen) > 1 else ""}


def split_qty(total: float, weights: list[float]) -> list[float]:
    """Нэг тоог ЖИНГЭЭР нь хуваана — нийлбэр нь ЯГ `total` (№88, 97).

    Блүүмийн 4,294ш нь `технологи 2,044 · архангай 326 · дарь эх 1,924` гэж
    гурван ТАЛБАЙД задарна; авлага нь НЭГ. Дугуйлалт нь нийлбэрийг эвдэж
    болохгүй тул үлдэгдлийг хамгийн том жинтэйд нь өгнө (largest remainder).
    """
    tw = sum(w for w in weights if w > 0)
    if total <= 0 or tw <= 0:
        return [0.0] * len(weights)
    raw = [(w / tw) * total if w > 0 else 0.0 for w in weights]
    base = [float(int(x)) for x in raw]
    left = round(total - sum(base))
    order = sorted(range(len(raw)), key=lambda i: (-(raw[i] - base[i]), -weights[i], i))
    i = 0
    while left > 0 and order:
        base[order[i % len(order)]] += 1
        left -= 1
        i += 1
    return base


# ── Захын тэмдэглэлийн үгсийн сан (№112) ────────────────────────────────────
NOTE_WORDS = ("тооцов", "ирээгүй", "хаав", "илүү", "нөат", "модонд", "кранд",
              "байршуулаагүй", "төлнө", "дутуу", "тооцоо дууссан", "суутга",
              "бартер", "шилжүүлсэн", "буцаалт", "алданга", "хугацаа хэтэр",
              "шивсэн", "шивэлцэнэ", "хасч тооцно", "барьцаанаас")
#: Тэмдэглэл БОЛОХГҮЙ (хүснэгтийн толгой, шошго).
NOTE_SKIP = ("нийт", "тоо", "үнэ", "хоног", "дүн", "гарсан", "ирсэн", "материал",
             "төлсөн", "үлдэгдэл", "өмнөх үлд", "акт", "барьцаа", "хугацаа",
             "харилцагч", "№")

YELLOW = "FFFFFF00"
TAN = "theme8"


def is_margin_note(text) -> bool:
    """Чөлөөт текст нүд нь ЗАХЫН ТЭМДЭГЛЭЛ мөн үү (шошго/толгой биш)."""
    s = _txt(text)
    if not s or len(s) > 220:
        return False
    low = s.casefold()
    if low in NOTE_SKIP:
        return False
    return any(w in low for w in NOTE_WORDS)


# ══════════════════════════════════════════════════ 2. ДЭВТЭР УНШИХ ДАВХРАГА

class Report:
    """Бүх алдаа, туг, шийдвэрийг НЭГ газар цуглуулна."""

    def __init__(self):
        self.warnings: list[str] = []
        self.unparsed: list[str] = []
        self.merges: list[dict] = []
        self.gaps: list[str] = []

    def warn(self, msg: str):
        if msg not in self.warnings:
            self.warnings.append(msg)

    def cannot_parse(self, where: str, value):
        self.unparsed.append(f"{where}: {value!r}")


def _rows(ws, max_col=None):
    return [tuple(r) for r in ws.iter_rows(max_col=max_col or ws.max_column,
                                           values_only=True)]


def _txt(c) -> str:
    return "" if c is None else str(c).strip()


# ── WB3 · AR самбар ────────────────────────────────────────────────────────
def parse_ar_board(rows, rep: Report, formulas=None, year: int | None = None) -> dict[str, dict]:
    """«Түрээс тооцоо-26»: [нэр, өмнөх, 3–7 сар, Нийт дүн, Тооцоо хийсэн,
    Үлдэгдэл, Барьцаа]. Нэргүй мөр + дүнтэй = хүснэгт дуусав.

    ⚠ ТЭМДГИЙГ ТОМЬЁОНООС уншина (№106): бүх мөр `«=+H-I»` (нэхэмжилсэн −
    төлсөн) атлаа Хурд группын `J32 «=+I32-H32»` ЭСРЭГ бичигдсэн — тэр
    78,165,000₮ нь ӨР биш, ИЛҮҮ ТӨЛӨЛТ.
    """
    formulas = formulas or {}
    header = rows[1] if len(rows) > 1 else ()
    out: dict[str, dict] = {}
    for ri, r in enumerate(rows[2:], start=3):
        r = tuple(r) + (None,) * 12
        name = _txt(r[0])
        if not name:
            if parse_money(r[1]) or parse_money(r[2]):
                break                                  # нийлбэрийн мөр
            continue
        total, paid = parse_money(r[7]), parse_money(r[8])
        raw_bal = parse_money(r[9])
        sign = sign_from_formula(formulas.get((ri, 10)))
        if raw_bal is not None and sign < 0:
            raw_bal = -raw_bal
            rep.warn(f"AR самбар «{name}» (J{ri}): үлдэгдлийн томьёо ЭСРЭГ "
                     f"«=+I{ri}-H{ri}» — {abs(raw_bal):,.0f}₮ нь ӨР БИШ, ИЛҮҮ "
                     f"ТӨЛӨЛТ (кредит) болгов")
        bal, dep, note = reconcile_balance_and_deposit(
            total, paid, raw_bal, parse_deposit(r[10]))
        not_lodged = isinstance(r[10], str) and "байршуул" in str(r[10]).casefold()
        c = canon_client(name)
        if not c.matched:
            rep.warn(f"AR самбар: нэр толинд алга — «{name}» (хэвээр нь авав)")
        row = out.setdefault(c.name, {"name": c.name, "sources": [], "total": 0.0,
                                      "paid": 0.0, "balance": 0.0, "deposit": 0.0,
                                      "deposit_not_lodged": False, "notes": [],
                                      "rows": 0, "last_covered": None})
        row["sources"].append(name)
        row["total"] += total or 0
        row["paid"] += paid or 0
        row["balance"] += bal
        row["deposit"] += dep
        row["deposit_not_lodged"] = row["deposit_not_lodged"] or not_lodged
        row["rows"] += 1
        # САМБАР хаана хүртэл нэхсэн бэ (§9-ийн зүүн тал). Нэг харилцагч
        # хоёр мөртэй бол ХОЙД нь — аль нь ч алдагдах учиргүй.
        lc = board_last_covered(header, r, year) if year else None
        if lc is not None and (row["last_covered"] is None
                               or lc.isoformat() > row["last_covered"]):
            row["last_covered"] = lc.isoformat()
        if note:
            row["notes"].append(note)
        if row["rows"] > 1:
            rep.merges.append({"canonical": c.name, "kind": "самбарын давхар мөр",
                               "sources": list(row["sources"]),
                               "detail": f"{row['rows']} мөр нийлүүлэв"})
    return out


def parse_ar_totals(rows, rep: Report) -> dict:
    """Самбарын НИЙЛБЭРИЙН мөр (нэргүй, эхний баганад дүнтэй).

    Түүний нийлбэр нь мөрүүдийнхээ нийлбэртэй ТААРАХ албагүй: SUM-ийн муж нь
    гараар бичигдсэн бөгөөд хүснэгтийн ЭХНИЙ мөрүүдийг тасалж орхисон байж
    болно. Тиймээс ХОЁУЛАНГ нь бичиж, зөрүүг нь тайланд гаргана."""
    for r in rows[2:]:
        r = tuple(r) + (None,) * 12
        if _txt(r[0]):
            continue
        if parse_money(r[1]) or parse_money(r[2]):
            return {"prev": parse_money(r[1]), "total": parse_money(r[7]),
                    "paid": parse_money(r[8]), "balance": parse_money(r[9]),
                    "deposit": parse_money(r[10])}
    rep.warn("AR самбар: нийлбэрийн мөр олдсонгүй")
    return {}


def parse_small_receivables(rows, rep: Report) -> list[dict]:
    """Самбарын доод «тооцов» жагсаалт — жижиг авлагууд, «Нийт»-д зогсоно."""
    out, started = [], False
    for r in rows:
        r = tuple(r) + (None,) * 4
        if not started:
            started = any("тооцов" in _txt(c) for c in r)
            continue
        name = _txt(r[0])
        if not name:
            continue
        if name == "Нийт":
            break
        amt = parse_money(r[1])
        if amt and amt > 0:
            out.append({"name": canon_client(name).name, "balance": round(amt),
                        "deposit": 0.0, "note": "Жижиг авлага («тооцов» жагсаалт)"})
    return out


def parse_loans(rows, rep: Report) -> list[dict]:
    """«Өглөгө.зээл-26»: [№, нэр, үндсэн, сарын хүү %, эхэлсэн огноо, …]"""
    out = []
    for r in rows:
        r = tuple(r) + (None,) * 6
        name = _txt(r[1])
        principal = parse_money(r[2])
        rate = parse_money(r[3])
        if not name or not principal or principal <= 0:
            continue
        start = parse_date_any(r[4])
        if r[4] not in (None, "") and start is None:
            rep.cannot_parse(f"Зээл «{name}» огноо", r[4])
        low = name.casefold()
        kind = ("bank" if "банк" in low else
                "credit" if "кредит" in low else "private")
        out.append({"name": name, "kind": kind, "principal": round(principal),
                    "monthly_rate": rate or 0,
                    "start_date": start.isoformat() if start else None,
                    "note": "" if rate else "Хүүгүй өглөг гэж бүртгэв"})
        if not rate:
            rep.warn(f"Зээл «{name}»: сарын хүү бичигдээгүй — 0%-иар авав")
        if "старкэйч" in low:
            break                       # доош нь хүснэгтийн гадуурх тэмдэглэл
    return out


# ── WB1 · тооллого ─────────────────────────────────────────────────────────
GROUP_PREFIX = {"хэв": "Хэв хашмал {}", "тулаас": "Тулаас {}", "труба": "Труба {}"}


def parse_stock(rows, rep: Report) -> list[dict]:
    """«тооллого 6.22»: [_, бүлэг, код, төлөв, тоо] — бүлэг/код доош өвлөнө.
    Төлөвгүй мөр = дэд нийлбэр (тухайн кодод өмнө нь төлөвтэй мөр орсон бол)."""
    acc: dict[tuple, float] = {}
    group = code = None
    for r in rows:
        r = tuple(r) + (None,) * 6
        g = _txt(r[1]).casefold()
        if g in GROUP_PREFIX:
            group = g
        if r[2] not in (None, ""):
            code = _txt(int(r[2]) if isinstance(r[2], float) and r[2] == int(r[2])
                        else r[2])
        state = _txt(r[3])
        cnt = parse_money(r[4])
        if group is None or code is None or cnt is None or cnt <= 0:
            continue
        if not state:
            if any(k[0] == group and k[1] == code for k in acc):
                continue                                       # дэд нийлбэр
            state = "А"
        acc[(group, code, state)] = acc.get((group, code, state), 0) + cnt
    out = []
    for (g, c, s), q in sorted(acc.items()):
        mat = material_of(c)
        if mat is None:
            rep.warn(f"Тооллого: каталогт байхгүй SKU — {GROUP_PREFIX[g].format(c)} "
                     f"({q:g}ш ачаалагдсангүй)")
            continue
        out.append({"material": GROUP_PREFIX[g].format(c), "grade": s, "on_hand": q})
    return out


#: Паркийн дэвтрийн ХАШААНЫ мөр («2026 шинэ» 51-р мөр).
YARD_ROW_NAME = "хашаанд бгаа"
YARD_REF = "2026 шинэ!51 (Хашаанд бгаа)"


def parse_park_yard(rows) -> dict[tuple[str, str], float]:
    """«2026 шинэ»-ийн `Хашаанд бгаа` мөр → {(материал, зэрэглэл): тоо}.

    Матрицын задлагч энэ мөр дээр ЗОГСДОГ (STOP) тул хашааны тоо ХААЯА Ч
    уншигддаггүй байв. Гэтэл «тооллого 6.22»-д трубаны тооны нүд БҮГД хоосон:
    хашааны мөр л 717/13/1,175/1,302/159/0/200-ыг мэднэ.
    """
    hdr = None
    for ri, r in enumerate(rows[:12]):
        if any(_txt(c) == "Харилцагч" for c in r):
            hdr = ri
            break
    if hdr is None:
        return {}
    code_row = tuple(rows[hdr + 1]) + (None,) * 4
    group_row = tuple(rows[hdr]) + (None,) * 4
    colmap: dict[int, tuple[str, str]] = {}
    for ci in range(2, len(code_row)):
        raw = _txt(code_row[ci])
        # `AE6` хоосон, `AE5='Шат'` — кодын мөргүй ГАНЦ баганат бүлэг. Түүнийг
        # алгасвал хашаанд байгаа 34 шат хаана ч бүртгэгдэхгүй өнгөрнө.
        if not raw:
            raw = _txt(group_row[ci])
        if not raw or raw.casefold() in ("нийт хэв", "нийт"):
            continue
        mat = material_of(raw)
        if mat is not None:
            colmap[ci] = mat
    for r in rows[hdr + 2:]:
        r = tuple(r) + (None,) * 4
        if _txt(r[1]).casefold() != YARD_ROW_NAME:
            continue
        out: dict[tuple[str, str], float] = {}
        for ci, mat in colmap.items():
            q = parse_money(r[ci]) if ci < len(r) else None
            if q:
                out[mat] = out.get(mat, 0.0) + float(q)
        return out
    return {}


def fill_stock_from_yard(stock: list[dict], yard: dict, rep: Report) -> list[dict]:
    """Тооллогод ХООСОН үлдсэн МАТЕРИАЛЫГ хашааны мөрөөр нөхнө.

    Дүрэм нь ЯГ НЭГ: тооллого тухайн материалын талаар ЮУ Ч хэлээгүй
    (нэг ч зэрэглэлээр мөр алга) бол — ба ЗӨВХӨН тэр үед — хашааны тоог
    зэрэглэл «А»-гаар нэмнэ. Тооллого нэг ч мөр бичсэн бол түүнийг ХӨНДӨХГҮЙ:
    түүний өөрийн тоолсон тоо ялна.

    Хоёр дэвтэр 4,215ш зөрдөг (63,695 ↔ 59,480) — энэ нөхөлт тэр зөрүүг
    ЗАСАХГҮЙ, зөвхөн ОГТ мэдэгдэхгүй байсан мөрүүдийг нээнэ. Зөрүү нь
    тайлангийн §9-д асуулт болж очно.
    """
    known = {r["material"] for r in stock}
    out = list(stock)
    by_mat: dict[str, float] = {}
    for (m, _g), q in yard.items():
        if m in known or q <= 0:
            continue
        by_mat[m] = by_mat.get(m, 0.0) + float(q)
    for m, q in sorted(by_mat.items()):
        out.append({"material": m, "grade": "А", "on_hand": q, "source": YARD_REF})
        rep.warn(f"Агуулах: «{m}» тооллогод хоосон — паркийн дэвтрийн хашааны "
                 f"мөрөөс {q:g}ш авав ({YARD_REF})")
    return out


# ── WB1 · харилцагчийн хуудас (циклийн блокууд) ────────────────────────────
_CYCLE_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\s*[-–]\s*(\d{1,2})\.(\d{1,2})")
_NO_RE = re.compile(r"№\s*(\d{1,2}\s*[./]\s*\d{1,3})")
_NO_FALLBACK = re.compile(r"№\s*(\d{1,3})")

# Хуудас → каноник харилцагч. Клиент биш хуудсууд: None + шалтгаан.
SHEET_CLIENT: dict[str, str | None] = {
    "ТэнгисТэмүүлэгч": None, "Марч-1": "Марч констракшн", "БЛҮҮМ-2": "Блүүм технологи",
    "Блүүт тооцоо": None, "Зулаа-3": "Зулаа", "Говьресурс-4": "Говь Ресурс Хурд",
    "говь тооцоо": None, "ГрэйтМайнинг-5": "Грэйт Майнинг", "Бэлгүтэй-6": "Бэлгүтэй",
    "Бутангууд-7": "Бутангууд", "Батцоож": None, "Бутан-Өнөорд": None,
    "ӨнөОрд-8": "Өнө Орд ХХК яармаг", "өнөорд-Архив": None, "Мөнхболд-9": "Мөнхболд",
    "Монуул-10": "Монуул ХХК", "Монуул тооцоо": None,
    "АшидДонж-11": "Ашид Донж Билгүүн", "БМ-12": "БМонгол ХХК",
    "Удам-13": "Удам констракшн", "ШижирТиара-14": "Шижир Тиара",
    "Агар Эрдэнэс-15": "Агар Эдэнэс Трэйд", "Арвинбулаг-16": "Арвинбулаг констракшн",
    "Өрнүүн эрчит зам-17": "Өрнүүн Эрчит Зам ххк", "Сүхжин-18": "Сүхжин ХХК",
    "Хатанбүүвэй-19": "Хатанбүүвэй", "Өнөорд Нүхт-20": "Өнө Орд ХХК нүхт",
    "Автокран": None, "Төвшөө": None, "тооллого 6.22": None,
    "Цэрэнчимэд-21": "Цэрэнчимэд", "НьюМэнПовэр-22": "Нью Мэн Повер ХХК",
    "Жаргалант ургамал-23": "Жаргалант Ургамал", "эрдэнэт шанд-24": "Эрдэнэт шанд",
    "Содоо худалдаа": None, "Зайсан үнийн санал": None,
}
NON_CLIENT_REASON = {
    "ТэнгисТэмүүлэгч": "худалдааны тооцоо (түрээс биш)",
    "Блүүт тооцоо": "Блүүмийн тооцоо нийлсэн хуудас",
    "говь тооцоо": "Говьресурсын тооцоо нийлсэн хуудас",
    "Батцоож": "Бутангуудын (№25/19) дэд дэвтэр — эзэн нь тодорхойгүй",
    "Бутан-Өнөорд": "харилцагч ХООРОНДЫН тооцоо (H11) — систем дэмжээгүй",
    "өнөорд-Архив": "архив",
    "Монуул тооцоо": "тооцоо нийлсэн хуудас",
    "Автокран": "кран — механизмын дэвтэр",
    "Төвшөө": "тэмдэглэл (эзэн нь тодорхойгүй)",
    "тооллого 6.22": "нөөцийн тооллого",
    "Содоо худалдаа": "худалдааны тооцоо",
    "Зайсан үнийн санал": "үнийн санал",
}


# Тоо/буцаалтын баганыг тэр ГУРВАН янзаар нэрлэдэг.
QTY_LABELS = ("тоо", "тоо ш", "тоо/ш")
# ⚠ E2: «буцаалт» бол ГУРАВ ДАХЬ нэршил (Марч-1!I7, Зулаа-3!I7/AS7). Урьд нь
# энэ багана танигдахгүй тул Зулаагийн буцаалтууд ГАДАА гэж тоологдох эрсдэлтэй
# байв.
BACK_LABELS = ("ирсэн", "орлого", "буцаалт", "буцсан")   # буцаж ирсэн огноо
OUT_LABELS = ("гарсан", "зарлага", "олгосон")            # олгосон огноо


def find_blocks(rows, max_header_row=14) -> list[dict]:
    """Хэвтээ блокийн толгойнууд. «үнэ»-г тулгуур болгоно, учир нь:
      * «тоо» ба «үнэ» хооронд хоосон багана байж болно (БЛҮҮМ-2 блок 2);
      * тооны баганыг «тоо» биш, ХАРИЛЦАГЧИЙН нэрээр нэрлэсэн байж болно
        (Өнөорд Нүхт-20 → «Өнөорд»);
      * буцаалтын багана «ирсэн» эсвэл «орлого» гэж нэрлэгддэг (Монуул-10).
    """
    blocks = []
    for ri, row in enumerate(rows[:max_header_row]):
        cells = [_txt(c).casefold() for c in row]
        for ci, c in enumerate(cells):
            if c != "үнэ" or ci == 0:
                continue
            days = next((j for j in range(ci + 1, min(ci + 4, len(cells)))
                         if cells[j] == "хоног"), None)
            if days is None:
                continue
            qty = next((j for j in range(max(ci - 3, 0), ci)
                        if cells[j] in QTY_LABELS), None)
            if qty is None:
                qty = ci - 1                      # «Өнөорд» гэх мэт нэршил
            cols = {"тоо": qty, "үнэ": ci, "хоног": days}
            for j in range(days + 1, min(days + 5, len(cells))):
                if cells[j] in BACK_LABELS:
                    cols["ирсэн"] = j
                elif cells[j] in OUT_LABELS:
                    cols["гарсан"] = j
            blocks.append({"header_row": ri, "cols": cols})
    return blocks


def block_period_end(rows, qty_col) -> tuple[int, int]:
    """Блокийн «Нийт» мөрний циклийн шошгоос үеийн ТӨГСГӨЛ (сар, өдөр)."""
    best = (0, 0)
    lo, hi = max(qty_col - 1, 0), qty_col + 5
    for r in rows:
        for c in tuple(r)[lo:hi]:
            m = _CYCLE_RE.search(_txt(c))
            if m:
                best = max(best, (int(m.group(3)), int(m.group(4))))
    return best


def _row_period(cells, qc) -> tuple[int, int] | None:
    """Мөрөнд циклийн шошго («4.30-5.31») байвал түүний ТӨГСГӨЛ."""
    for c in cells[max(qc - 3, 0):qc + 5]:
        m = _CYCLE_RE.search(_txt(c))
        if m:
            return (int(m.group(3)), int(m.group(4)))
    return None


def _resolve_code(cells, qc, carry):
    """SKU-гийн нүд блок бүрд өөр байрлана:
      «Материал | код | тоо»  (Грэйт)  → qc-1
      «код | тэмдэглэл | тоо» (Агар)   → qc-2, дунд нь «үлд» гэсэн тэмдэглэл
      хоосон                            → өмнөх мөрөө өвлөнө
      «шинэ»/«пл»                       → өмнөх SKU + ЭНЭ зэрэглэл
      «пластик» + `6012` ЗЭРЭГЦЭЭ      → 6012-ийн ПЛАСТИК зэрэглэл (E3)
    Буцна: ((материал, зэрэглэл) | None, түүхий утга, өвлүүлэх шинэ carry)"""
    prim = cells[qc - 1] if qc >= 1 else None
    sec = cells[qc - 2] if qc >= 2 else None
    # E3: зэрэглэлийг НЭРИЙН баганад бичсэн — «пластик»/«шинэ»/«плас»
    word = grade_token(prim) or grade_token(sec)
    for raw in (prim, sec):
        got = material_of(raw)
        if got:
            if got[1] == "А" and word:
                got = (got[0], word)
            return got, raw, got
    if word and carry:
        return (carry[0], word), prim, carry
    if _txt(prim) == "" and _txt(sec) == "" and carry:
        return carry, None, carry
    return None, (prim if _txt(prim) else sec), carry


def parse_contract_sheet(sheet_name, rows, rep: Report, formulas=None, fills=None,
                         year: int | None = None):
    """Хуудасны ХАМГИЙН СҮҮЛИЙН блокийн ХАМГИЙН СҮҮЛИЙН үеэс одоо ГАДАА
    байгаа мөрүүд.

    Түүний хуудас циклээ ХОЁР янзаар давхарлана: баруун тийш шинэ блок
    (БЛҮҮМ) БА доош шинэ үе (Удам). Хоёуланг нь уншина — эс тэгвэл Удамын
    7.06-нд бүгдийг буцаасныг олж харалгүй 1,600ш-ийг гадаа гэж бүртгэнэ.

    ⚠ E1 — ХЭН ГАДАА БАЙГААГ ТЭР ӨӨРӨӨ ХЭЛНЭ. Үеийн «Нийт» тооны нүд нь
    томьёогоороо АЛЬ МӨРҮҮДИЙГ тоолж байгаагаа зөөж явдаг:

        Бутангууд-7!V61 «=+V58+V55+…+V10»   ← 16 мөр СОНГОСОН
        ГрэйтМайнинг-5!D22 «=SUM(D7:D21)»   ← бүх мөр

    Тэр жагсаалтад ОРООГҮЙ мөр бол буцаалт (эсвэл тэр өөрөө тоолохоо больсон)
    — ГАДАА БИШ. Урьд нь задлагч зөвхөн «ирсэн» баганыг хардаг байсан тул
    ОГНООГҮЙ буцаалт бүрийг гадаа гэж тоолж, Бутангуудыг 4,432ш болгосон
    (түүний өөрийн тоо 1,879).
    """
    formulas = formulas or {}
    fills = fills or {}
    blocks = find_blocks(rows)
    if not blocks:
        return None, {"error": "тоо·үнэ·хоног толгой олдсонгүй"}
    block = max(blocks, key=lambda b: block_period_end(rows, b["cols"]["тоо"]))
    cols = block["cols"]
    qc, rc = cols["тоо"], cols["үнэ"]
    dc = cols.get("хоног")
    irsen = cols.get("ирсэн")

    # ---- блокийг босоо үеүүдэд хуваах (мөрийн ИНДЕКСЭЭ хамт зөөнө) ----
    sections: list[dict] = []
    cur: list[tuple[int, tuple]] = []
    labels: list[str] = []
    for ri in range(block["header_row"] + 1, len(rows)):
        cells = tuple(rows[ri]) + (None,) * 12
        if any(_txt(c).startswith("Нийт төлөх") for c in cells[max(qc - 3, 0):qc + 2]):
            break
        per = _row_period(cells, qc)
        if per:
            for c in cells[max(qc - 3, 0):qc + 5]:
                if _CYCLE_RE.search(_txt(c)):
                    labels.append(_txt(c))
            sections.append({"period": per, "rows": cur, "total_row": ri})
            cur = []
            continue
        cur.append((ri, cells))
    if cur:
        sections.append({"period": (0, 0), "rows": cur, "total_row": None})

    def _live_rows(sec):
        return sum(1 for _, cells in sec["rows"]
                   if parse_money(cells[qc]) and parse_money(cells[rc]))

    filled = [s for s in sections if _live_rows(s)]
    head = _header_text(rows)
    hdr = parse_contract_header(head)
    # ХАМРАЛТ (P0-11): дэвтэр ХААНА хүртэл нэхэгдсэн бэ — тооцоо ЯГ маргааш
    # нь эхэлнэ (`billing_from`). Жилгүй бол тооцохгүй (задлагчийн нэгж тест).
    lc = last_covered_from_labels(labels, year) if year else None
    # ХУУДАСНЫ ӨӨРИЙН ҮЛДЭГДЭЛ (2-р шат) — OB нь эндээс гарна, самбараас БИШ:
    # ингэснээр үлдэгдэл ба хамралт хоёр НЭГ дэвтрээс гарна.
    bal = sheet_balance(rows, qc, (dc if dc is not None else rc) + 1, sheet_name)
    base_meta = {"no": hdr["no"] or _contract_no(rows), "header": head,
                 "start_date": hdr["date"].isoformat() if hdr["date"] else None,
                 "appendix": hdr["appendix"],
                 "cycle_mode": cycle_mode_from_labels(labels),
                 "cycle_labels": labels, "blocks": len(blocks),
                 "last_covered": lc.isoformat() if lc else None,
                 "balance": bal["amount"] if bal else None,
                 "balance_ref": bal["ref"] if bal else "",
                 "sections": len(sections)}
    if not filled:
        return [], {**base_meta, "vat_mentioned": _vat_seen(rows),
                    "returned_rows": 0, "period_end": (0, 0), "empty": True,
                    "her_total": None, "total_ref": ""}
    section = max(filled, key=lambda s: s["period"])

    # ---- ТҮҮНИЙ «Нийт» тоо: утга + аль мөрүүдийг тоолсон ----
    her_total = None
    total_ref = ""
    counted: set[int] | None = None
    tr = section.get("total_row")
    if tr is not None:
        her_total = parse_money((tuple(rows[tr]) + (None,) * 12)[qc])
        total_ref = f"{_col_name(qc)}{tr + 1}"
        picked = parse_total_formula(formulas.get((tr + 1, qc + 1)))
        if picked:
            counted = {r - 1 for r in picked["rows"]}     # 1-based → 0-based
            if picked["kind"] == "refs":
                rep.warn(f"{sheet_name}: «Нийт» тоо нь ГАРААР ТҮҮСЭН "
                         f"({total_ref} = {len(picked['rows'])} мөр) — тэр мөрүүдийг л "
                         f"гадаа гэж авав")

    items: dict[tuple, float] = {}
    rate_refs: dict[tuple, str] = {}
    carry = None
    skipped_returned = 0
    dropped: list[str] = []
    for ri, cells in section["rows"]:
        code, raw_code, carry = _resolve_code(cells, qc, carry)
        qty, rate = parse_money(cells[qc]), parse_money(cells[rc])
        if qty is None or rate is None:
            continue
        if not (50 <= rate <= 2000) or not (0 < qty <= 50000):
            continue
        # ---- E1: буцаалт мөр үү? ----
        back = _txt(cells[irsen]) if irsen is not None else ""
        returned = bool(back) and "ирээгүй" not in back.casefold()
        if counted is not None and ri not in counted:
            returned = True                       # түүний Нийт үүнийг тоолоогүй
        elif counted is None:
            fill = fills.get((ri + 1, qc + 1), "")
            days = parse_money(cells[dc]) if dc is not None else None
            if fill.startswith(TAN) or (days is not None and days <= 0):
                returned = True
        if returned:
            skipped_returned += 1
            continue
        if code is None:
            dropped.append(f"{_txt(raw_code)} ({qty:g}ш)")
            rep.warn(f"{sheet_name}: SKU таних боломжгүй — «{_txt(raw_code)}» "
                     f"({qty:g}ш × {rate:g}₮ алгасав)"
                     + (f" · {CATALOG_GAPS[_txt(raw_code)]}"
                        if _txt(raw_code) in CATALOG_GAPS else ""))
            continue
        items[(code[0], code[1], rate)] = items.get((code[0], code[1], rate), 0) + qty
        # ТАРИФЫН НҮД — каталогийн `base_rate` ЭНД цитатлагдана (2-р шат):
        # «330₮ гэж хаанаас гарав» гэсэн асуултад ЭНЭ мөр хариулна.
        rate_refs.setdefault((code[0], code[1], rate),
                             f"{sheet_name}!{_col_name(rc)}{ri + 1}")

    meta = {**base_meta, "vat_mentioned": _vat_seen(rows),
            "returned_rows": skipped_returned, "period_end": section["period"],
            "her_total": her_total, "total_ref": total_ref, "dropped": dropped}
    out = [{"material": m, "grade": g, "qty": q, "daily_rate": rt,
            "rate_ref": rate_refs.get((m, g, rt), "")}
           for (m, g, rt), q in sorted(items.items())]
    return out, meta


def sheet_balance(rows, qty_col: int, amount_col: int, sheet: str) -> dict | None:
    """ТҮҮНИЙ ХУУДАСНЫ ӨӨРИЙН «Үлдэгдэл» нүд — эхний үлдэгдлийн эх сурвалж.

    `БЛҮҮМ-2!X33 = X30 − X32` = 382,179,050₮. Самбар нь тэр харилцагчийг
    392,791,500₮ гэдэг: хоёр дэвтэр 10,612,450₮-өөр зөрж байна. Үлдэгдлийг
    ХУУДСААС нь авбал үлдэгдэл БА хамралт (`last_covered`) хоёр НЭГ дэвтрээс
    гарах тул `billing_from = хамралт + 1` нь бүтцээрээ зөв болно; самбараас
    авбал хоёр өөр дэвтрийн тоо нийлж, залгаа нь ЯМАР Ч баталгаагүй.

    Долоон хуудсанд ДӨРВӨН нэршил: «Үлдэгдэл», «үлдэгдэл», «Үлдэгдэл төлбөр»,
    «Нийт төлөх үлдэгдэл дүн/төлбөр». Тиймээс «үлд» гэсэн ХЭСГЭЭР хайна.
    «Өмнөх үлдэгдэл» бол ЭХЛЭЛ — сонгогдохгүй. Ижил үг нэг блокт хоёр удаа
    гарвал (`W30` ба `T33`) ДООД талынх нь ялна: тэр нь эцсийн мөр.
    """
    best = None
    for ri, r in enumerate(rows, start=1):
        cells = tuple(r) + (None,) * (amount_col + 2)
        for ci in range(max(qty_col - 4, 0), amount_col + 1):
            t = _txt(cells[ci]).casefold().strip(" :")
            if "үлд" not in t or "өмнөх" in t:
                continue
            v = parse_money(cells[amount_col])
            if v is None:
                continue
            best = {"amount": float(v),
                    "ref": f"{sheet}!{_col_name(amount_col)}{ri}"}
            break
    return best


def _col_name(ci: int) -> str:
    """0-ээс тоологдох баганын индекс → Excel-ийн үсэг (A, B, … AA)."""
    s, n = "", ci + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _header_text(rows) -> str:
    """Гэрээний толгойн мөр — эхний 3 мөрийн ХАМГИЙН УРТ «№…» агуулсан текст.

    Зулаа-3-д толгой нь 2-р мөр, Марч-1-д 1-р мөр, Бутангууд-д C1/L1/U1
    гурвуулаа ижил.
    """
    best = ""
    for hr in rows[:3]:
        for c in tuple(hr):
            s = _txt(c)
            if "№" in s and "хавсралт" in s.casefold() and len(s) > len(best):
                best = s
    if best:
        return best
    for hr in rows[:3]:
        for c in tuple(hr):
            s = _txt(c)
            if "№" in s and len(s) > len(best):
                best = s
    return best


def _contract_no(rows) -> str:
    no = ""
    for hr in rows[:3]:
        head = " ".join(_txt(c) for c in tuple(hr)[:6])
        m = _NO_RE.search(head) or _NO_FALLBACK.search(head)
        if m:
            no = re.sub(r"\s+", "", m.group(1))
            break
    return no


def _vat_seen(rows) -> bool:
    return any("нөат" in _txt(c).casefold() for r in rows for c in r)


# ── WB2 · паркийн матриц ───────────────────────────────────────────────────
def parse_park_matrix(rows, rep: Report, sheet_label: str):
    """«2026 шинэ»: МӨР=харилцагч × БАГАНА=SKU. Хоёр толгой мөр (бүлэг, код).
    Буцна: {каноник: {(материал, зэрэглэл): тоо}} + танигдаагүй баганууд."""
    hdr = None
    for ri, r in enumerate(rows[:12]):
        if any(_txt(c) == "Харилцагч" for c in r):
            hdr = ri
            break
    if hdr is None:
        rep.warn(f"{sheet_label}: «Харилцагч» толгой олдсонгүй — матриц уншаагүй")
        return {}, []
    code_row = tuple(rows[hdr + 1]) + (None,) * 4
    group_row = tuple(rows[hdr]) + (None,) * 4
    colmap: dict[int, tuple[str, str]] = {}
    unknown: list[str] = []
    for ci in range(2, len(code_row)):
        raw = _txt(code_row[ci])
        if not raw or raw.casefold() in ("нийт хэв", "нийт"):
            continue
        # «V2 шинэ» — зэрэглэл баганын гарчигт шингэсэн
        mat = material_of(raw)
        if mat is None:
            grp = _txt(group_row[ci])
            unknown.append(f"{grp + ' ' if grp else ''}{raw}".strip())
            continue
        colmap[ci] = mat
    STOP = {"түрээсийн дүн", "хашаанд бгаа", "үндсэн материал", "худалдаж авсан",
            "бнсу орж ирсэн", "хятад улсаас ирсэн", "зарагдсан", "акт",
            "акт орж ирсэн", "аа-д", "тооцоолуур", "түрээсэнд", "үлдэгдэл"}
    out: dict[str, dict] = {}
    raw_out: dict[str, dict] = {}
    for r in rows[hdr + 2:]:
        r = tuple(r) + (None,) * 4
        name = _txt(r[1])
        if not name:
            continue
        if name.casefold() in STOP:
            break
        c = canon_client(name)
        bucket = out.setdefault(c.name, {})
        # ⚠ ТҮҮХИЙ нэрээр НЭГ ДАХЬ хуулбар: `БЛҮҮМ технологи` нь каноник
        # `Блүүм технологи` руу нийлдэг тул ТАЛБАЙН задаргаа (№97) каноник
        # хүснэгтээс ОЛДОХГҮЙ болно.
        raw_bucket = raw_out.setdefault(name, {})
        for ci, mat in colmap.items():
            q = parse_money(r[ci]) if ci < len(r) else None
            if q:
                bucket[mat] = bucket.get(mat, 0) + q
                raw_bucket[mat] = raw_bucket.get(mat, 0) + q
        if not c.matched:
            rep.warn(f"Паркийн дэвтэр: нэр толинд алга — «{name}» (хэвээр нь авав)")
    return out, sorted(set(unknown)), raw_out


# ══════════════════════════════════ 2b. ЗУРГААН ШИНЭ БАЙР (H8/H11/№69/72/112)

class Book:
    """Нэг дэвтрийн ГУРВАН давхарга: утга · томьёо · дүүргэлт/фонт.

    Өнгө нь тэр дэвтэрт ӨГҮҮЛБЭР (№111): бор-шаргал = буцаагдсан багц,
    ШАР = «анхаар», улаан фонт = нэхэгдэх/маргаантай. Урьд нь зөвхөн утгын
    давхарга уншигддаг байсан тул эдгээр бүгд унадаг байв.
    """

    def __init__(self, path):
        import openpyxl
        self.v = openpyxl.load_workbook(path, data_only=True)
        self.f = openpyxl.load_workbook(path, data_only=False)

    @property
    def sheetnames(self):
        return [ws.title.strip() for ws in self.v.worksheets]

    def rows(self, sheet, max_col=None):
        return _rows(self.v[sheet], max_col=max_col)

    def formulas(self, sheet) -> dict[tuple[int, int], str]:
        """{(мөр, багана) 1-ээс: томьёоны текст} — томьёотой нүд л орно."""
        out = {}
        for row in self.f[sheet].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    out[(c.row, c.column)] = c.value
        return out

    def styles(self, sheet) -> tuple[dict, dict]:
        """({(мөр, багана): дүүргэлт}, {(мөр, багана): фонтын өнгө})."""
        fills, fonts = {}, {}
        for row in self.v[sheet].iter_rows():
            for c in row:
                if c.value is None:
                    continue
                fl = c.fill
                if fl is not None and fl.fill_type is not None and fl.fgColor is not None:
                    fg = fl.fgColor
                    if fg.type == "rgb" and fg.rgb not in (None, "00000000"):
                        fills[(c.row, c.column)] = str(fg.rgb)
                    elif fg.type == "theme":
                        fills[(c.row, c.column)] = f"theme{fg.theme}"
                col = getattr(c.font, "color", None)
                if col is not None and col.type == "rgb" and \
                        col.rgb not in (None, "FF000000", "00000000"):
                    fonts[(c.row, c.column)] = str(col.rgb)
        return fills, fonts


def harvest_deposit(rows, formulas, as_of, sheet: str) -> dict:
    """Барьцааны ГҮЙДЭГ ДЭВТЭР → явдлууд (H8 / R22).

    `'байршуулаагүй'` бол 0 БИШ — ЯВДАЛ ОГТ БОЛООГҮЙ (№55): `status='none'`,
    явдал ХООСОН. Гинж бол гишүүн бүр нэг явдал.
    """
    out = {"events": [], "status": "", "ref": "", "amount": None}
    for ri, r in enumerate(rows, start=1):
        cells = tuple(r) + (None,) * 12
        for ci, c in enumerate(cells, start=1):
            if "барьцаа" not in _txt(c).casefold():
                continue
            for cc in range(ci + 1, min(ci + 12, len(cells) + 1)):
                raw = cells[cc - 1]
                if raw is None:
                    continue
                if isinstance(raw, str) and "байршуул" in raw.casefold():
                    return {"events": [], "status": "none",
                            "ref": f"{sheet}!{_col_name(cc - 1)}{ri}", "amount": None}
                amt = parse_money(raw)
                if amt is None or amt <= 0:
                    continue
                ref = f"{sheet}!{_col_name(cc - 1)}{ri}"
                terms = parse_money_chain(formulas.get((ri, cc))) or [amt]
                return {"events": deposit_events_from_chain(terms, as_of, ref),
                        "status": "held", "ref": ref, "amount": amt}
    return out


#: Хүний нэр БОЛОХГҮЙ үгс (хүснэгтийн шошго гарын үсгийн бүсэд ч тааралдана).
NOT_A_NAME = {"нийт", "төлсөн", "акт", "барьцаа", "үлд", "үлдэгдэл", "дүн", "тоо",
              "үнэ", "хоног", "гарсан", "ирсэн", "буцаалт", "материал", "хэв",
              "тулаас", "труба", "булан", "шинэ", "пластик", "плас", "зээл",
              "бартер", "илүү", "төлбөр", "хугацаа", "өмнөх", "мод", "панэр",
              "прүүс", "шида", "банз", "пин", "тросс", "чүлдий", "кран"}


def harvest_contacts(rows, sheet: str) -> list[dict]:
    """Гарын үсгийн блок → ХАРИЛЦАГЧИЙН ХҮМҮҮС (№72, 73).

    ЗӨВХӨН түрээслэгчийн тал: «Жигүүр Зам ХХК» болон Отгонцэцэг/Батмөнх нь
    БИДНИЙ гарын үсэг, харилцагчийн хүн БИШ.
    """
    OURS = ("жигүүр зам", "отгонцэцэг", "батмөнх", "тооцоо нийлсэн",
            "тооцоог гаргасан", "тооцоог нэгтгэсэн", "тооцоог нийлсэн")
    # Гарын үсгийн БҮС — эхний «Тооцоо нийлсэн / түрээслэгч» мөрөөс доош.
    # Түүнээс дээш 8 оронтой тоо бол УТАС биш, МӨНГӨ (36,883,800₮).
    sig_from = len(rows) + 1
    for ri, r in enumerate(rows, start=1):
        if any(("тооцоо" in _txt(c).casefold() and "нийлсэн" in _txt(c).casefold())
               or _txt(c).casefold().startswith(("түрээслэгч", "түрээлэгч"))
               for c in tuple(r)):
            sig_from = ri
            break
    out: list[dict] = []
    seen: set[str] = set()
    for ri, r in enumerate(rows, start=1):
        cells = tuple(r) + (None,) * 12
        for ci, c in enumerate(cells, start=1):
            s = _txt(c)
            if not s or len(s) < 4 or len(s) > 120:
                continue
            low = s.casefold()
            if any(o in low for o in OURS):
                continue
            extra = [cells[j] for j in range(ci, min(ci + 6, len(cells)))]
            # «Гантөмөр» 99125325 88432001 — цэггүй, шошгогүй, гэхдээ УТАСТАЙ
            bare = (ri >= sig_from
                    and low not in NOT_A_NAME
                    and re.fullmatch(r"[^\W\d_][^\W\d_.\s]{1,29}", s)
                    and any(isinstance(x, int) and not isinstance(x, bool)
                            and 60_000_000 <= x <= 99_999_999 for x in extra))
            if not bare:
                if not (any(w in low for w in ROLE_WORDS) or "…" in s or "...." in s):
                    continue
                if not re.search(r"[.…]{3,}", s) and ":" not in s:
                    continue
            role_only = (not bare and ":" in s
                         and re.fullmatch(r"\s*(?:[^\W\d_]|\s)+\s*:?\s*", s))
            text = s
            if role_only:                       # «Нярав :» | «Н.Соль» | 99966285
                nxt = next((_txt(x) for x in extra if _txt(x)), "")
                if not nxt:
                    continue
                text = f"{s.strip()} {nxt}"
                if any(o in text.casefold() for o in OURS):
                    continue
            got = parse_signatory(text, extra)
            if not got or not got["name"]:
                continue
            if re.fullmatch(r"[\W\d_]+", got["name"]):
                continue
            # ХУУЛИЙН ЭТГЭЭД бол ХҮН биш — «түрээслэгч: Бутангууд … ХХК» нь
            # блокийн толгой, гарын үсэгтэн биш.
            if _LEGAL.search(got["name"].casefold()) or len(got["name"].split()) > 4:
                continue
            key = name_key(got["name"])
            if not key or key in seen:
                continue
            seen.add(key)
            got["ref"] = f"{sheet}!{_col_name(ci - 1)}{ri}"
            out.append(got)
    return out


#: ХАМГИЙН БАГА УТГАТАЙ ТЭМДЭГЛЭЛ — ГУРВАН ҮГ. «буцаалт», «модонд», «хаав» нь
#: өгүүлбэр биш, ХҮСНЭГТИЙН шошго: гэрээний хажууд наагаад тэр уншихад
#: юу ч хэлэхгүй. Гурван үгнээс эхлээд л ТҮҮНИЙ өгүүлбэр болно
#: («барьцаанаас суутгаж тооцов», «хугацаа хэтэрвэл … алданга тооцно»).
MEMO_MIN_WORDS = 3

#: ТЭМДЭГЛЭЛИЙН ЗОХИОГЧ — хоёрхон нэр. «Шилжүүлэлт» гэдэг нь ХЭРЭГСЛИЙН нэр
#: байсан: тэр хүнийг мэдэхгүй. Дэвтрээс уншсныг «Дэвтрээс», системийн
#: өөрийн асуултыг «Систем» гэж бичнэ.
NOTE_AUTHOR_BOOK = "Дэвтрээс"
NOTE_AUTHOR_SYSTEM = "Систем"


def harvest_notes(rows, formulas, fills, fonts, sheet: str,
                  as_of) -> tuple[list[dict], list[dict]]:
    """ЗАХЫН ТЭМДЭГЛЭЛ (№111, 112 / P1-22) — ТҮҮНИЙХ ба МАШИНЫХ хоёр овоолго.

    Буцаана `(kept, dropped)`:

    · `kept` — ТҮҮНИЙ бичсэн, гурав ба түүнээс дээш үгтэй захын тэмдэглэл.
      Текст нь ЯГ ХЭВЭЭРЭЭ: нүдний хаяг (`Зулаа-3!AQ27`) НААГДАХГҮЙ, туг
      өргөгдөхгүй. Тэр өөрийнхөө өгүүлбэрийг гэрээн дээрээ таньж харна.

    · `dropped` — МАШИНЫ хэл: «ШАР нүд: 35,000,000 — «бартер»», «МАРГААНТАЙ
      гар тоо», «улаанаар бичсэн дүн», ганц үгтэй шошго. Эдгээр нь ХӨГЖҮҮЛЭГЧИД
      хэрэгтэй мөр — тайлангийн хавсралтад очно, гэрээн дээр НААГДАХГҮЙ.

    Урьд нь гурвуулаа НЭГ жагсаалтаар гэрээ рүү ордог байсан тул тэр өөрийн
    дэвтрээ нээхэд «ШАР нүд: … · Блүүт тооцоо!G8» гэсэн 215 мөр угтдаг байв.
    """
    kept: list[dict] = []
    dropped: list[dict] = []
    seen: set[str] = set()
    for ri, r in enumerate(rows, start=1):
        cells = tuple(r) + (None,) * 12
        for ci, c in enumerate(cells, start=1):
            if c is None:
                continue
            ref = f"{sheet}!{_col_name(ci - 1)}{ri}"
            fill = fills.get((ri, ci), "")
            font = fonts.get((ri, ci), "")
            yellow = fill == YELLOW
            text = _txt(c)
            numeric = isinstance(c, (int, float)) and not isinstance(c, bool)
            near = next((_txt(cells[j]) for j in range(max(ci - 4, 0), ci - 1)
                         if _txt(cells[j]) and not _txt(cells[j]).isdigit()), "")
            body, kind = "", ""
            if is_margin_note(text):
                body, kind = text, "захын тэмдэглэл"
            elif yellow and text:
                body = (f"ШАР нүд: {c:,.0f}" if numeric else f"ШАР нүд: {text}")
                if numeric and not near:
                    continue                      # тайлагдахгүй ганц тоо
                body += f" — «{near}»" if near else ""
                kind = "шар нүд"
            elif font == "FFFF0000" and numeric and abs(c) >= 1e6:
                # ГАРААР бичсэн улаан тоо = МАРГААНТАЙ нэхэмжлэл;
                # томьёотой улаан тоо = ердийн «нэхэгдэх» нийлбэр (баримт).
                hand = (ri, ci) not in formulas
                body = (("МАРГААНТАЙ гар тоо " if hand else "улаанаар бичсэн дүн ")
                        + f"{c:,.0f}₮" + (f" — «{near}»" if near else ""))
                kind = "улаан дүн"
            if not body:
                continue
            key = f"{body}|{ref}"
            if key in seen:
                continue
            seen.add(key)
            if kind != "захын тэмдэглэл" or len(body.split()) < MEMO_MIN_WORDS:
                dropped.append({"text": body, "ref": ref, "kind": kind,
                                "sheet": sheet})
                continue
            day = parse_date_any(text, default_year=as_of.year)
            if day is None:
                m = re.match(r"^\s*(\d{1,2})[.](\d{1,2})", text)
                if m:
                    day = parse_date_any(f"{m.group(1)}.{m.group(2)}",
                                         default_year=as_of.year)
            # Нүдэнд орсон давхар зай нь дэлгэц дээр эмх замбараагүй харагдана;
            # үг нь хэвээрээ.
            kept.append({"text": " ".join(body.split()), "date": str(day or as_of),
                         "flag": False, "author": NOTE_AUTHOR_BOOK, "ref": ref})
    return kept, dropped


def dedupe_notes(notes: list[dict], dropped: list[dict], client: str,
                 where: str) -> list[dict]:
    """НЭГ ижил өгүүлбэрийг НЭГ Л удаа наана.

    Гэрээний нөхцөлийн мөр («хугацаа хэтэрвэл … алданга тооцно») хуудсанд
    цикл бүрийн хажууд ДАХИН ДАХИН бичигдсэн байдаг: нүдний хаяг нь өөр тул
    урьд нь найман хуулбар болж очдог байв. Хаяг нь текстээс гарсны дараа
    тэдгээр нь ЯГ ижил мөрүүд — үлдсэн нь хавсралт руу.
    """
    def key_of(n):
        # «НӨАТ: «…»» нь ТҮҮНИЙ ижил өгүүлбэрийг ТУГТАЙГААР давтдаг — хоёулаа
        # наагдвал нэг зүйлийг хоёр удаа асууна. Бүрхүүлийг нь хуулж
        # түлхүүрийг тэнцүүлнэ; ТУГТАЙ нь ялна (доорх эрэмбэ).
        s = " ".join(str(n.get("text", "")).split())
        m = re.fullmatch(r"НӨАТ:\s*«(.*)»", s)
        return " ".join((m.group(1) if m else s).split()).casefold()

    out, seen = [], set()
    for n in sorted(notes, key=lambda x: not x.get("flag")):
        key = key_of(n)
        if not key or key in seen:
            dropped.append({"client": client, "where": where,
                            "kind": "давхардсан", "text": n.get("text", ""),
                            "ref": n.get("ref", ""), "sheet": ""})
            continue
        seen.add(key)
        out.append(n)
    return out


AE_BLOCK = "А Е блок"
OTHER_SITE = "Бусад талбай"


def harvest_ae_block(rows) -> dict[str, dict]:
    """`Батцоож!F5='А Е блок үлдэгдэл'` — ДЭД-ОБЪЕКТЫН per-SKU үлдэгдэл (№88).

    Нэг гэрээний ДОТОРХ талбайн задаргаа: `F7 «=+B7-C7-D7-E7»`. Буцаана:
    {талбай: {(материал, зэрэглэл): тоо}} — `_site_split`-ийн жин.
    """
    hdr = fc = sc = None
    for ri, r in enumerate(rows[:12]):
        cells = tuple(r) + (None,) * 4
        for ci, c in enumerate(cells):
            if AE_BLOCK.casefold() in _txt(c).casefold():
                hdr, fc = ri, ci
            elif _txt(c).casefold() == "үлдэгдэл" and hdr is not None and sc is None:
                sc = ci
    if hdr is None or fc is None:
        return {}
    out: dict[str, dict] = {AE_BLOCK: {}, OTHER_SITE: {}}
    for r in rows[hdr + 1:]:
        cells = tuple(r) + (None,) * 4
        mat = material_of(cells[0])
        if mat is None:
            continue
        f = parse_money(cells[fc]) or 0
        s = (parse_money(cells[sc]) if sc is not None else None) or 0
        if f <= 0 and s <= 0:
            continue
        out[AE_BLOCK][mat] = out[AE_BLOCK].get(mat, 0) + max(f, 0)
        out[OTHER_SITE][mat] = out[OTHER_SITE].get(mat, 0) + max(s - f, 0)
    return out if any(out[AE_BLOCK].values()) else {}


def ae_block_snapshot(rows, year: int) -> str:
    """`Батцоож!F` баганын ЗУРАГЛАСАН ОГНОО — «А Е блок»-оос ЗҮҮН талын
    сүүлчийн огнооны толгой (`E6='4.01'` → 2026-04-01).

    Энэ багана нь `B−C−D−E` буюу ТЭР ӨДРИЙН байдал: цаашид 4.08, 4.15, 4.21-нд
    юу гарсан/ирснийг тоолоогүй. Тиймээс түүгээр өнөөдрийн 1,879ш-ийг хуваах
    нь ХУГАЦААГАА ӨНГӨРСӨН харьцаагаар хуваах явдал.
    """
    hdr = fc = None
    for ri, r in enumerate(rows[:12]):
        for ci, c in enumerate(tuple(r) + (None,) * 4):
            if AE_BLOCK.casefold() in _txt(c).casefold():
                hdr, fc = ri, ci
    if hdr is None or hdr + 1 >= len(rows):
        return ""
    best = ""
    for ci, c in enumerate(tuple(rows[hdr + 1])):
        if ci >= fc:
            break
        d = parse_date_any(_txt(c), default_year=year)
        if d is not None:
            best = d.isoformat()
    return best


def harvest_reg(rows) -> str:
    """`'РД:'` шошгын хажуугийн регистрийн дугаар (№6, ГрэйтМайнинг-5!C5)."""
    for r in rows[:12]:
        cells = tuple(r) + (None,) * 6
        for ci, c in enumerate(cells):
            if re.fullmatch(r"\s*рд\s*:?\s*", _txt(c), re.IGNORECASE):
                for cc in range(ci + 1, min(ci + 4, len(cells))):
                    v = cells[cc]
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        return str(int(v))
                    if _txt(v).isdigit():
                        return _txt(v)
    return ""


def harvest_agreed(rows, sheet: str, as_of) -> dict | None:
    """«Тооцоо нийлсэн» + ОГНОО → нэхэмжлэлийн ХАМТАРСАН ТӨЛӨВ (№69 / P1-5).

    `Блүүт тооцоо!D2` «Түрээсийн төлбөр тооцоо нийлсэн баримт» + `H4` =
    `'2026.07.20'` — ХОЁР ТАЛ гарын үсгээр баталсан гэдэг нь ТӨЛӨВ.
    """
    hit = False
    for r in rows[:8]:
        for c in tuple(r):
            if "тооцоо" in _txt(c).casefold() and "нийлсэн" in _txt(c).casefold():
                hit = True
    if not hit:
        return None
    best = None
    for ri, r in enumerate(rows[:8], start=1):
        for ci, c in enumerate(tuple(r), start=1):
            s = _txt(c)
            if "№" in s or "хавсралт" in s.casefold():
                continue                      # энэ бол ГЭРЭЭНИЙ толгой, тооцоо биш
            d = parse_date_any(c)
            if d and 2020 <= d.year <= 2100 and (best is None or d > best[0]):
                best = (d, f"{sheet}!{_col_name(ci - 1)}{ri}")
    if best is None:
        return None
    return {"date": str(best[0]), "by": f"{sheet} — тооцоо нийлсэн баримт",
            "ref": best[1]}


# ═══════════════════════════════════════════════════════════ 3. УГСРАЛТ

#: Аудитын §1.2-ийн ЭРЭМБЭ (Нийт дүн, tie-break Үлдэгдэл). `--clients top10`.
TOP10 = ["Хурд групп", "Бутангууд", "Блүүм технологи", "Ашид Донж Билгүүн",
         "Грэйт Майнинг", "Өнө Орд ХХК яармаг", "Голден лайт",
         "Марч констракшн", "Дархан Оюунаа", "Зулаа"]

#: Топ-10-ын ХАРИЛЦАГЧ БҮРИЙН нэмэлт хуудсууд (statement, матриц, хоорондын).
EXTRA_SHEETS: dict[str, list[str]] = {
    "Блүүм технологи": ["Блүүт тооцоо"],
    "Бутангууд": ["Батцоож", "Бутан-Өнөорд"],
    "Өнө Орд ХХК яармаг": ["өнөорд-Архив"],
}

#: Талбайн задаргаа — WB2 паркийн дэвтрийн МӨР бүр нэг ТАЛБАЙ (№88, 97).
SITES: dict[str, list[str]] = {
    "Блүүм технологи": ["БЛҮҮМ технологи", "БЛҮҮМ архангай", "Блүүм дарь эх"],
}

#: Гэрээний ДАНСАН дээр байгаа боловч ТҮРЭЭС БИШ мөрүүд → тэмдэглэл (H11).
#: (Самбарын `Үлдэгдэл` эдгээрийг АЛЬ ХЭДИЙН агуулсан тул бичилт үүсгэхгүй.)
ACCOUNT_NOTES: dict[str, list[tuple[str, float, str]]] = {
    "Бутангууд": [
        ("2025 онд бэлэн мөнгө зээлсэн нийт дүн", 164_492_000, "Бутан-Өнөорд!G23"),
        ("Төгрөгийн зээл (2026-04-08, Батцоож)", 10_000_000, "Бутан-Өнөорд!G24"),
        ("Төгрөгийн зээл (данс, Батцоож)", 3_000_000, "Бутан-Өнөорд!G25"),
        ("Авто кран түрээс — 5 удаа × 500,000₮", 2_500_000, "Бутан-Өнөорд!G26"),
        ("Шланз машин түрээс — 5 удаа × 300,000₮", 1_500_000, "Бутан-Өнөорд!G27"),
        ("Бутангууд ХХК-ын ажилчдын цалин", 2_800_000, "Бутан-Өнөорд!G28"),
        ("Дутагдуулсан бараа материалын акт (5 SKU)", 34_029_000, "Батцоож!AC19"),
        ("Бэлэн мөнгөний актын бараа (10 мөр)", 40_970_500, "Бутангууд-7!AH18"),
    ],
    "Ашид Донж Билгүүн": [
        ("Авто кран түрээс 6.1-6.30", 10_000_000, "АшидДонж-11!P30"),
        ("Авто кран төлсөн 6.23", 8_000_000, "АшидДонж-11!P31"),
        ("Авто кран үлдэгдэл", 2_000_000, "АшидДонж-11!P32"),
        ("Авто кран түрээс 7.01-7.31", 10_000_000, "АшидДонж-11!Y46"),
        ("Авто кран түрээс 8.01-8.31", 10_000_000, "АшидДонж-11!AH35"),
        ("Байрны төлбөр — 2026-07-20 «төлнө»", 170_000_000, "АшидДонж-11!Y52"),
        ("Байрны төлбөр СУУТГАВ", 170_000_000, "АшидДонж-11!AH50"),
        ("Авлага", 3_000_000, "АшидДонж-11!AH39"),
    ],
    "Хурд групп": [
        ("Бартер-байрны хэлцэл: 95.86м² × 5,600,000₮", 536_816_000, "2026 тооцоо!G38"),
        ("Санал-бодит зөрүү («алдагдал»)", 28_758_000, "2026 тооцоо!J38"),
        ("НӨАТ 10% суурь дүнгээс", 148_830_000, "2026 тооцоо!I36"),
        ("Үндсэн хөрөнгийн бүртгэлийн авлага (2026-04-09)", 1_028_000_000,
         "Үндсэн хөрөнгө 26!P20"),
    ],
    "Голден лайт": [
        ("Үндсэн хөрөнгийн бүртгэлд авлага ХЭВЭЭР (самбар 0 гэж)", 117_566_060,
         "Үндсэн хөрөнгө 26!P17"),
    ],
    "Дархан Оюунаа": [
        ("Үндсэн хөрөнгийн бүртгэлд «Дарханчулуун» нэрээр, «модонд» төлнө",
         59_400_000, "Үндсэн хөрөнгө 26!P22"),
    ],
}

#: ХАРИЛЦАГЧ ХООРОНДЫН цэвэр тооцоо — САМБАРЫН R24 нь Бутангуудын `Үлдэгдэл`
#: мөрөнд ОРООГҮЙ тул энэ нь ЦОРЫН ГАНЦ нэмэлт мөнгө (H11 / №81, 85).
NETTING: list[dict] = [
    {"client": "Бутангууд", "kind": "transfer", "amount": 139_648_000,
     "label": "Өнө Ордтой тооцоо — 2026.06.22 акт",
     "ref": "2026 тооцоо!R24 · Бутан-Өнөорд",
     "decisions": [("Бутан-Өнөорд!G32", 124_648_000, "хоёр талын цэвэр тооцоо"),
                   ("Батцоож!AC23 · Бутан-Өнөорд!G33", 142_351_500, "хоёр хуудсанд"),
                   ("Батцоож!AC22", 459_981_564, "нийт төлөх (самбартай 15,000,000₮ зөрөв)")]},
]


def _grade_blind_weight(bucket: dict, material: str) -> float:
    """Зэрэглэлээ ҮЛ ХАРААД тухайн МАТЕРИАЛЫН бүх тоог нэмнэ.

    Паркийн дэвтрийн толгойд `Y6='V2'` ба `Z6='V2 шинэ'` гэсэн ХОЁР багана
    байдаг тул `Z18`-ийн 600ш нь `Тулаас В2·шинэ` болж задардаг; гэрээний
    мөр нь `Тулаас В2·А`. Хоёр түлхүүр таарахгүй тул дарь эхийн жин 0 болж,
    600ш нь технологи руу нүүж 2,585/385/1,324 гэсэн ХУДАЛ задаргаа гардаг байв.
    """
    return float(sum(q for (m, _g), q in bucket.items() if m == material))


def _site_split(client: str, items: list[dict], park: dict, rep: Report,
                explicit: dict | None = None,
                snapshot: str = "") -> tuple[list[dict], dict | None]:
    """Нэг гэрээ, ОЛОН ТАЛБАЙ (№88, 97) — олголтыг талбай тус бүрд хуваана.

    Блүүмийн хуудас ГУРВАН талбайг барина (`технологи 2,044 · архангай 326 ·
    дарь эх 1,924` = 4,294ш) атлаа АВЛАГА нь НЭГ. Хуваалт нь паркийн дэвтрийн
    SKU-гийн тоогоор жинлэгдэж, нийлбэр нь хуудасны тоотой ЯГ тэнцэнэ.

    Буцаана `(sites, problem)`. `problem` нь §9-д очих АСУУЛТ: эх сурвалж нь
    ОГНООТОЙ ЗУРАГ бөгөөд нийлбэр нь өнөөдрийн тоотой таарахгүй бол систем
    ХУВААХГҮЙ — нэг хоосон нүднээс төрсөн харьцааг «түүний хуваарилалт» гэж
    бичих нь ЗОХИОМОЛ баримт болно (Батцоож!F27 = 11,866ш, гэрээ нь 1,879ш).
    """
    if explicit:
        park, names = explicit, list(explicit)
    else:
        names = SITES.get(client)
    if not names:
        return [], None
    want = sum(i["qty"] for i in items)
    if explicit:
        src_total = sum(sum(d.values()) for d in explicit.values())
        if abs(src_total - want) > max(1.0, 0.01 * want):
            text = (f"«{AE_BLOCK}» / «{OTHER_SITE}»-ийн хуваалт: эх сурвалж "
                    + (f"{snapshot}-ний зураг" if snapshot else "огнооны зураг")
                    + f" ({src_total:,.0f}ш), одоогийн {want:,.0f}ш-тэй "
                      f"таарахгүй — та хуваарилна уу")
            rep.warn(f"«{client}»: {text}")
            return [], {"client": client, "text": text, "source": src_total,
                        "now": want, "snapshot": snapshot}
    # SKU тус бүрд паркийн дэвтрийн тоо жин болно; ЯГ таарах (материал,
    # зэрэглэл) байхгүй бол ЗЭРЭГЛЭЛГҮЙГЭЭР нэмнэ; тэр ч байхгүй бол ТАЛБАЙН
    # НИЙТ тоо (2,044 · 326 · 1,924) руу уналт хийнэ — «бүгдийг эхний талбайд»
    # гэж хаях нь худал задаргаа болно.
    site_totals = [float(sum(park.get(n, {}).values())) for n in names]
    out = [{"site": n, "items": []} for n in names]
    for it in items:
        key = (it["material"], it["grade"])
        w = [float(park.get(n, {}).get(key, 0) or 0)
             or _grade_blind_weight(park.get(n, {}), it["material"])
             for n in names]
        if not any(w):
            w = site_totals if any(site_totals) else [1.0] + [0.0] * (len(names) - 1)
        for i, q in enumerate(split_qty(float(it["qty"]), w)):
            if q > 0:
                out[i]["items"].append({**it, "qty": q})
    got = sum(i["qty"] for s in out for i in s["items"])
    if abs(got - want) > 0.5:                    # хуваалт нийлбэрээ эвдсэн бол ХАЯНА
        rep.warn(f"«{client}»: талбайн хуваалт нийлбэрээ эвдэв "
                 f"({got:g} ≠ {want:g}) — хуваалгүй нэг мөрөөр авав")
        return [], None
    rep.warn(f"«{client}»: {want:g}ш нь {len(names)} ТАЛБАЙД задарлаа — "
             + " · ".join(f"{s['site']} {sum(i['qty'] for i in s['items']):g}ш"
                          for s in out))
    return [s for s in out if s["items"]], None


def build(src_dir: str, as_of: str, clients_mode: str = "top10",
          ob_source: str = "sheet") -> tuple[dict, Report]:
    rep = Report()
    as_of_d = date.fromisoformat(as_of)
    dropped_notes: list[dict] = []
    site_gaps: list[dict] = []

    def load(fn):
        path = os.path.join(src_dir, fn)
        if not os.path.exists(path):
            raise SystemExit(f"Файл олдсонгүй: {path}")
        return Book(path)

    b1, b2, b3 = load(WB1), load(WB2), load(WB3)
    wb1, wb2, wb3 = b1.v, b2.v, b3.v

    # ---- WB3: самбар, жижиг авлага, зээл ----
    board_rows = _rows(wb3["Түрээс тооцоо-26"], max_col=14)
    board = parse_ar_board(board_rows, rep, b3.formulas("Түрээс тооцоо-26"),
                           year=as_of_d.year)
    board_totals = parse_ar_totals(board_rows, rep)
    small = parse_small_receivables(board_rows, rep)
    loans = parse_loans(_rows(wb3["Өглөгө.зээл-26"], max_col=12), rep)

    # ---- WB1: тооллого ----
    stock = parse_stock(_rows(wb1["тооллого 6.22"], max_col=10), rep)

    # ---- WB2: паркийн матриц (гадаа байгаа) ----
    park_rows = _rows(wb2["2026 шинэ"])
    yard = parse_park_yard(park_rows)
    # Тооллогод ХООСОН үлдсэн материалыг (бүх труба, Тулаас В6, Шат) хашааны
    # мөрөөр нөхнө — эс тэгвэл 7 труба SKU бүгд 0-ээр ачаалагдана.
    stock = fill_stock_from_yard(stock, yard, rep)
    park, unknown_sku, park_raw = parse_park_matrix(park_rows, rep,
                                                    "2026 шинэ")
    park_prev, _, _ = parse_park_matrix(_rows(wb2["2026 он"]), rep, "2026 он")
    for u in unknown_sku:
        code = u.split()[-1]
        rep.warn(f"Паркийн дэвтэр: каталогт байхгүй багана — {u}"
                 + (f" ({CATALOG_GAPS[code]})" if code in CATALOG_GAPS else ""))
        rep.gaps.append(u)

    # ---- WB1: харилцагчийн хуудсууд ----
    sheets_seen = {ws.title.strip() for ws in wb1.worksheets}
    for name in sorted(sheets_seen - set(SHEET_CLIENT)):
        rep.warn(f"WB1: тодорхойлогдоогүй хуудас — «{name}» (алгасав)")

    contracts: list[dict] = []
    sheet_report: list[dict] = []
    used_no: dict[str, str] = {}
    sheet_contacts: dict[str, list[dict]] = {}
    sheet_agreed: dict[str, dict] = {}
    sheet_reg: dict[str, str] = {}
    #: харилцагч → {"amount", "ref", "last_covered", "sheet"} — OB-ийн эх сурвалж
    sheet_ob: dict[str, dict] = {}
    for sheet, client in SHEET_CLIENT.items():
        if sheet not in sheets_seen:
            rep.warn(f"WB1: хуудас алга — «{sheet}»")
            continue
        if client is None:
            sheet_report.append({"sheet": sheet, "client": None,
                                 "reason": NON_CLIENT_REASON.get(sheet, "клиент биш")})
            continue
        if clients_mode == "top10" and client not in TOP10:
            sheet_report.append({"sheet": sheet, "client": client, "filtered": True,
                                 "wb1_qty": 0, "wb2_qty": 0, "wb2_listed": False,
                                 "result": "--clients top10 шүүлтүүрээс гадуур",
                                 "reason": "--clients top10 шүүлтүүрээс гадуур"})
            continue
        rows = _rows(wb1[sheet])
        f_sheet = b1.formulas(sheet)
        fills, fonts = b1.styles(sheet)
        items, meta = parse_contract_sheet(sheet, rows, rep, f_sheet, fills,
                                           year=as_of_d.year)
        in_park = client in park
        on_rent = park.get(client, {})
        out_qty = sum(on_rent.values())
        entry = {"sheet": sheet, "client": client, "meta": meta,
                 "wb1_qty": sum(i["qty"] for i in items) if items else 0,
                 "wb2_qty": out_qty, "wb2_listed": in_park}
        if items is None:
            rep.warn(f"{sheet} ({client}): {meta['error']} — гэрээ үүсээгүй")
            entry["result"] = "толгой олдсонгүй"
            sheet_report.append(entry)
            continue
        if not items:
            entry["result"] = "хуудасны сүүлийн үед гадаа мөр үлдээгүй"
            if out_qty > 0:
                rep.warn(f"{sheet} ({client}): хуудсанд гадаа мөр алга, харин "
                         f"паркийн дэвтэрт {out_qty:g}ш байна — ШИЙДВЭР ХЭРЭГТЭЙ")
            sheet_report.append(entry)
            continue
        # ХОЁР дэвтрийн санал зөрөх гурван тохиолдол (олдвор б):
        #   бүртгэлтэй & >0  → хоёул зөвшөөрөв   → гэрээ үүснэ
        #   бүртгэлгүй       → парк мэдэхгүй (8-р сарын шинэ гэрээ) → үүснэ + туг
        #   бүртгэлтэй & =0  → парк «буцсан» гэж БАТАЛЖ байна → ҮҮСГЭХГҮЙ + туг
        if in_park and out_qty <= 0:
            entry["result"] = "паркийн дэвтэр 0 гэж БАТАЛСАН — гэрээ үүсгээгүй"
            rep.warn(f"{sheet} ({client}): хуудсанд {entry['wb1_qty']:g}ш гадаа "
                     f"боловч «2026 шинэ» паркийн дэвтэр 0 гэж бичсэн — гэрээ "
                     f"ҮҮСГЭЭГҮЙ. АЛЬ нь үнэн бэ?")
            sheet_report.append(entry)
            continue
        if not in_park:
            rep.warn(f"{sheet} ({client}): паркийн дэвтэрт мөр АЛГА (8-р сарын "
                     f"шинэ гэрээ бололтой) — хуудсаар нь {entry['wb1_qty']:g}ш "
                     f"гадаа гэж үүсгэв, ТЭР баталгаажуулна")
        no = meta["no"] or sheet
        if no in used_no:
            rep.warn(f"Гэрээний дугаар давхардав: №{no} — «{used_no[no]}» ба "
                     f"«{sheet}». Хоёр дахийг №{no}·{sheet} болгов")
            no = f"{no}·{sheet}"
        used_no[no] = sheet

        # ---- НӨАТ: хуудас БА түүний statement хуудсууд (Блүүт тооцоо!G9) ----
        vat_sheets = [sheet] + [s for s in EXTRA_SHEETS.get(client, [])
                                if s in sheets_seen]
        vat_texts = [c for s in vat_sheets for r in _rows(wb1[s]) for c in r
                     if isinstance(c, str)]
        vat_pct, vat_notes = vat_of(vat_texts)

        # ---- Зургаан байр: барьцаа · холбоо · тэмдэглэл · тооцоо нийлсэн ----
        dep = harvest_deposit(rows, f_sheet, as_of_d, sheet)
        contacts = harvest_contacts(rows, sheet)
        notes, dropped = harvest_notes(rows, f_sheet, fills, fonts, sheet, as_of_d)
        dropped_notes += [dict(d, client=client, where="гэрээ") for d in dropped]
        agreed = None
        for extra in EXTRA_SHEETS.get(client, []):
            if extra not in sheets_seen:
                continue
            e_rows = _rows(wb1[extra])
            e_f = b1.formulas(extra)
            e_fill, e_font = b1.styles(extra)
            contacts += [c for c in harvest_contacts(e_rows, extra)
                         if name_key(c["name"]) not in
                         {name_key(x["name"]) for x in contacts}]
            e_keep, e_drop = harvest_notes(e_rows, e_f, e_fill, e_font, extra, as_of_d)
            notes += e_keep
            dropped_notes += [dict(d, client=client, where="гэрээ") for d in e_drop]
            agreed = agreed or harvest_agreed(e_rows, extra, as_of_d)
            if dep["status"] == "":
                dep = harvest_deposit(e_rows, e_f, as_of_d, extra)
        agreed = agreed or harvest_agreed(rows, sheet, as_of_d)
        # ТҮРЭЭС БИШ мөр · ГЭРЭЭНИЙ ТОЛГОЙ — ХӨГЖҮҮЛЭГЧИЙН мөр, ТҮҮНИЙХ БИШ.
        # Түүний гэрээн дээр «(самбарын Үлдэгдэлд АЛЬ ХЭДИЙН орсон, бичилт
        # үүсгээгүй)» гэсэн өгүүлбэр наах нь ямар ч асуулт тавихгүй, зөвхөн
        # чимээ болно: хавсралтад очно.
        for label, amount, ref in ACCOUNT_NOTES.get(client, []):
            dropped_notes.append({
                "client": client, "where": "гэрээ", "kind": "түрээс биш мөр",
                "text": f"{label} — {amount:,.0f}₮", "ref": ref, "sheet": sheet})
        for vn in vat_notes:
            notes.append({"text": "НӨАТ: «" + " ".join(str(vn).split()) + "»",
                          "date": str(as_of_d), "flag": True,
                          "author": NOTE_AUTHOR_BOOK, "ref": sheet})
        if meta.get("start_date"):
            dropped_notes.append({
                "client": client, "where": "гэрээ", "kind": "гэрээний толгой",
                "text": f"«{meta['header']}» — эхлэл {meta['start_date']}"
                        + (f", хавсралт №{meta['appendix']}"
                           if meta["appendix"] else ""),
                "ref": f"{sheet}!C1", "sheet": sheet})
        else:
            rep.warn(f"{sheet} ({client}): гэрээний толгойгоос ОГНОО олдсонгүй — "
                     f"эхлэл {as_of} болов")
        if meta["cycle_mode"] == "month":
            rep.warn(f"{sheet} ({client}): циклийн шошго нь КАЛЕНДАРЬ САР "
                     f"({', '.join(meta['cycle_labels'][:3])}) — cycle_mode=month")
        her = meta.get("her_total")
        loaded = sum(i["qty"] for i in items)
        if her is not None and abs(her - loaded) > 0.5:
            rep.warn(f"{sheet} ({client}): ТҮҮНИЙ Нийт {her:g}ш ≠ ачаалсан "
                     f"{loaded:g}ш ({meta['total_ref']}) — зөрүү {her - loaded:g}ш")

        # ---- ТАЛБАЙН задаргаа (№88, 97) ----
        ae, ae_snap = {}, ""
        if "Батцоож" in EXTRA_SHEETS.get(client, []) and "Батцоож" in sheets_seen:
            bt_rows = _rows(wb1["Батцоож"])
            ae = harvest_ae_block(bt_rows)
            ae_snap = ae_block_snapshot(bt_rows, as_of_d.year) if ae else ""
            if ae:
                dropped_notes.append({
                    "client": client, "where": "гэрээ", "kind": "талбайн хуваалт",
                    "text": f"«{AE_BLOCK}» нь Батцоож!F («{AE_BLOCK} үлдэгдэл» = "
                            f"B−C−D−E) баганаас жинлэгдэв — ДУНДЫН байдал",
                    "ref": "Батцоож!F5", "sheet": "Батцоож"})
        sites, site_problem = _site_split(client, items, park_raw, rep,
                                          explicit=ae or None, snapshot=ae_snap)
        if site_problem:
            site_gaps.append(dict(site_problem, sheet=sheet, no=meta.get("no") or sheet))

        # ---- ХАМРАЛТ: дэвтэр хаана дуусав → систем хаанаас эхлэх вэ (P0-11) ----
        last_covered = meta.get("last_covered")
        board_lc = (board.get(client) or {}).get("last_covered")
        if not last_covered:
            rep.warn(f"{sheet} ({client}): циклийн шошго олдсонгүй — дэвтэр "
                     f"ХЭЗЭЭ хүртэл нэхэгдсэн нь мэдэгдэхгүй тул тооцоо "
                     f"{as_of}-оос эхэлнэ. ТЭР баталгаажуулна")
        elif board_lc and board_lc != last_covered:
            rep.warn(f"{sheet} ({client}): самбар {board_lc} хүртэл, хуудас "
                     f"{last_covered} хүртэл нэхсэн — хамралт ЗӨРЖ байна "
                     f"(систем хуудсыг сонгов, §9-д бичив)")

        day = sum(i["qty"] * i["daily_rate"] for i in items)
        contracts.append({
            "client": client, "no": no, "items": items,
            "vat_percent": vat_pct,
            "start_date": meta.get("start_date"),
            # `last_covered` + 1 хоног = `billing_from` (ачаалагч тавина).
            # `board_last_covered` нь ЗӨВХӨН тулгалтын §9-д хэрэглэгдэнэ.
            "last_covered": last_covered,
            "board_last_covered": board_lc,
            "cycle_mode": meta["cycle_mode"],
            "sheet": sheet,
            "deposit_events": dep["events"],
            "deposit_status": dep["status"] or "none",
            "sites": sites,
            "notes": dedupe_notes(notes, dropped_notes, client, "гэрээ"),
            # ГЭРЭЭНИЙ ТЭМДЭГЛЭЛ нь ТҮҮНИЙ талбар — шилжүүлэлтийн машины
            # мөрийг («Шилжүүлэлт: «БЛҮҮМ-2» хуудсаас · өдрийн дүн …») тэнд
            # бичихгүй. Тэр өгөгдөл нь `audit`-д ба тулгалтын тайланд байна.
            "note": "",
        })
        sheet_ob[client] = {"amount": meta.get("balance"),
                            "ref": meta.get("balance_ref", ""),
                            "last_covered": last_covered, "sheet": sheet}
        sheet_contacts[client] = sheet_contacts.get(client, []) + contacts
        if agreed:
            sheet_agreed[client] = agreed
        reg = harvest_reg(rows)
        if reg:
            sheet_reg[client] = reg
        entry["result"] = "гэрээ үүсэв"
        entry["no"] = no
        entry["day_amount"] = day
        entry["her_total"] = her
        entry["total_ref"] = meta.get("total_ref", "")
        entry["returned_rows"] = meta.get("returned_rows", 0)
        entry["start_date"] = meta.get("start_date")
        entry["last_covered"] = last_covered
        entry["board_last_covered"] = board_lc
        entry["cycle_mode"] = meta["cycle_mode"]
        entry["vat_percent"] = vat_pct
        entry["deposit_events"] = len(dep["events"])
        entry["notes"] = len(notes)
        entry["contacts"] = len(contacts)
        entry["sheet_balance"] = meta.get("balance")
        entry["balance_ref"] = meta.get("balance_ref", "")
        entry["day_amount"] = day
        sheet_report.append(entry)

    # ---- Харилцагчдын жагсаалт ----
    clients: list[dict] = []
    seen: set[str] = set()
    keep = set(TOP10) if clients_mode == "top10" else None
    with_contract = {c["client"] for c in contracts}
    for name, row in board.items():
        if keep is not None and name not in keep:
            continue
        # САМБАРЫН захын үг нь ТҮҮНИЙХ — гурав ба түүнээс дээш үгтэй бол
        # тэмдэглэл болно, богино шошго нь хавсралт руу.
        c_notes = []
        for t in row["notes"]:
            if len(_txt(t).split()) >= MEMO_MIN_WORDS:
                c_notes.append({"text": _txt(t), "date": str(as_of_d),
                                "flag": False, "author": NOTE_AUTHOR_BOOK,
                                "ref": "Түрээс тооцоо-26"})
            else:
                dropped_notes.append({"client": name, "where": "харилцагч",
                                      "kind": "самбарын шошго", "text": _txt(t),
                                      "ref": "Түрээс тооцоо-26",
                                      "sheet": "Түрээс тооцоо-26"})
        entries = [dict(n, date=str(as_of_d)) for n in NETTING if n["client"] == name]
        for n in entries:
            for ref, val, why in n.get("decisions", []):
                dropped_notes.append({
                    "client": name, "where": "харилцагч", "kind": "шийдвэр",
                    "text": f"«{n['label']}»-ийн ӨӨР утга: {val:,.0f}₮ ({why})",
                    "ref": ref, "sheet": "2026 тооцоо"})
        # WB1-д ХУУДАСГҮЙ топ-10 (Хурд групп · Голден лайт · Дархан Оюунаа) —
        # гэрээ үүсэхгүй тул тэдний баримтжсан мөрүүд ХАРИЛЦАГЧ дээр буудаг.
        if name not in with_contract:
            for label, amount, ref in ACCOUNT_NOTES.get(name, []):
                dropped_notes.append({
                    "client": name, "where": "харилцагч", "kind": "түрээс биш мөр",
                    "text": f"{label} — {amount:,.0f}₮", "ref": ref,
                    "sheet": "2026 тооцоо"})
        contacts = sheet_contacts.get(name, [])
        ob = sheet_ob.get(name) or {}
        board_bal = round(row["balance"])
        sheet_bal = None if ob.get("amount") is None else round(ob["amount"])
        use_sheet = ob_source == "sheet" and sheet_bal is not None
        clients.append({"name": name,
                        "balance": sheet_bal if use_sheet else board_bal,
                        # ЭХНИЙ ҮЛДЭГДЭЛ ХААНААС ГАРАВ (2-р шат). Хуудастай
                        # харилцагчийн үлдэгдэл нь ТҮҮНИЙ ХУУДСААС гарна:
                        # тэгвэл үлдэгдэл ба хамралт (`ob_date`) НЭГ дэвтрээс
                        # гарах тул `billing_from = хамралт + 1` нь бүтцээрээ
                        # зөв болно. Самбар нь §9-д АСУУЛТ болж үлдэнэ.
                        "balance_source": "sheet" if use_sheet else "board",
                        "balance_sheet": sheet_bal,
                        "balance_board": board_bal,
                        "balance_ref": (ob.get("ref") if use_sheet
                                        else "Түрээс тооцоо-26!J"),
                        "ob_date": ob.get("last_covered"),
                        "deposit": round(row["deposit"]), "note": "",
                        "deposit_not_lodged": bool(row["deposit_not_lodged"]),
                        "reg": sheet_reg.get(name, ""),
                        "person": contacts[0]["name"] if contacts else "",
                        "phone": contacts[0]["phone"] if contacts else "",
                        "contacts": contacts,
                        "entries": [{k: v for k, v in e.items() if k != "decisions"}
                                    for e in entries],
                        "notes": dedupe_notes(c_notes, dropped_notes, name,
                                              "харилцагч"),
                        "agreed": sheet_agreed.get(name)})
        seen.add(name)
    for row in small:
        if keep is not None and row["name"] not in keep:
            continue
        if row["name"] in seen:
            rep.warn(f"«{row['name']}» самбар БА «тооцов» жагсаалтад давхар — "
                     f"жижиг авлагыг алгасав")
            continue
        clients.append(row)
        seen.add(row["name"])
    # Гэрээтэй боловч самбарт байхгүй харилцагч — 0 үлдэгдлээр нээнэ
    for c in contracts:
        if c["client"] not in seen:
            clients.append({"name": c["client"], "balance": 0, "deposit": 0,
                            "note": "", "balance_source": "none",
                            "balance_board": None, "balance_sheet": None,
                            "balance_ref": "", "ob_date": c.get("last_covered")})
            seen.add(c["client"])
            rep.warn(f"«{c['client']}»: AR самбарт мөр алга, гэрээ нь бий — "
                     f"үлдэгдэл 0-ээр нээв")
    # Паркийн дэвтэрт гадаа байгаа боловч хаана ч алга
    for name, mats in sorted(park.items()):
        if sum(mats.values()) > 0 and name not in seen:
            rep.warn(f"«{name}»: паркийн дэвтэрт {sum(mats.values()):g}ш гадаа "
                     f"байна, AR самбарт мөр алга — ШИЙДВЭР ХЭРЭГТЭЙ")

    alias_map = {v: canon for canon, vs in ALIASES.items() for v in vs}
    catalog = [{"name": n, **{k: v for k, v in spec.items() if k != "why"},
                "note": spec["why"]} for n, spec in sorted(CATALOG_NEW.items())]
    data = {
        "as_of": as_of,
        "clients_mode": clients_mode,
        # ЭХНИЙ ҮЛДЭГДЛИЙН ЭХ СУРВАЛЖ: `sheet` = харилцагчийн ӨӨРИЙН хуудас
        # (хуудасгүй бол самбар), `board` = урьдын зан төлөв.
        "ob_source": ob_source,
        "source": f"{WB1} + {WB2} + {WB3} (openpyxl, 2026.08 байдлаар)",
        "clients": sorted(clients, key=lambda c: c["name"]),
        "catalog": catalog,
        "stock": stock,
        "loans": loans,
        "barter": BARTER,
        "contracts": sorted(contracts, key=lambda c: c["no"]),
        "audit": {
            "top10": TOP10,
            "catalog_new": catalog,
            "decisions": [
                {"client": n["client"], "what": n["label"],
                 "a": n["amount"], "a_ref": n["ref"], "b": val, "b_ref": ref,
                 "why": why}
                for n in NETTING for ref, val, why in n.get("decisions", [])],
            "alias_map": dict(sorted(alias_map.items())),
            "ambiguous": [{"a": a, "b": b, "question": q} for a, b, q in AMBIGUOUS],
            "merges": rep.merges,
            "ar_board": {k: {kk: vv for kk, vv in v.items() if kk != "notes"}
                         for k, v in sorted(board.items())},
            "ar_board_totals": board_totals,
            "park_now": {k: {f"{m}·{g}": q for (m, g), q in sorted(v.items())}
                         for k, v in sorted(park.items()) if v},
            "park_prev": {k: {f"{m}·{g}": q for (m, g), q in sorted(v.items())}
                          for k, v in sorted(park_prev.items()) if v},
            "sheets": sheet_report,
            "catalog_gaps": sorted(set(rep.gaps)),
            # ХАЯГДСАН ТЭМДЭГЛЭЛ — гэрээн дээр НААГДААГҮЙ, гэхдээ АЛДАГДААГҮЙ.
            # Тайлангийн хавсралтад бүтнээрээ хэвлэгдэнэ (хөгжүүлэгчид).
            "dropped_notes": dropped_notes,
            "site_gaps": site_gaps,
            "park_yard": {f"{m}·{g}": q for (m, g), q in sorted(yard.items())},
            "contact_refs": sorted(
                {f"{c['name']} · {c.get('ref', '')}"
                 for lst in sheet_contacts.values() for c in lst if c.get("ref")}),
            "warnings": rep.warnings,
            "unparsed": rep.unparsed,
        },
    }
    return data, rep


def main(argv=None):
    ap = argparse.ArgumentParser(description="xlsx → real_data.json")
    ap.add_argument("--src", default=os.path.expanduser(
        "~/Downloads/Түрээсийн тооцоо 2026 2"), help="гурван .xlsx байгаа хавтас")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(
        os.path.abspath(__file__)), "real_data.json"))
    ap.add_argument("--as-of", default=DEFAULT_AS_OF)
    ap.add_argument("--clients", choices=("top10", "all"), default="top10",
                    help="top10 = аудитын §1.2-ийн эрэмбийн эхний арав (анхны)")
    ap.add_argument("--ob-source", choices=("sheet", "board"), default="sheet",
                    dest="ob_source",
                    help="эхний үлдэгдэл ХААНААС: sheet = харилцагчийн өөрийн "
                         "хуудас (анхны), board = «Түрээс тооцоо-26» самбар")
    a = ap.parse_args(argv)

    data, rep = build(a.src, a.as_of, a.clients, a.ob_source)
    with open(a.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1, sort_keys=False)
        f.write("\n")

    debt = sum(c["balance"] for c in data["clients"] if c["balance"] > 0)
    cred = -sum(c["balance"] for c in data["clients"] if c["balance"] < 0)
    from_sheet = sum(1 for c in data["clients"] if c.get("balance_source") == "sheet")
    print(f"→ {a.out}")
    print(f"  OB эх сурв: --ob-source {a.ob_source}  ·  хуудсаар {from_sheet} "
          f"харилцагч, самбараар {len(data['clients']) - from_sheet}")
    print(f"  Харилцагч : {len(data['clients'])}  ·  авлага {debt:,.0f}₮  ·  "
          f"илүү төлөлт {cred:,.0f}₮")
    print(f"  Барьцаа   : {sum(c['deposit'] for c in data['clients']):,.0f}₮")
    print(f"  Гэрээ     : {len(data['contracts'])}  ·  өдрийн нийт дүн "
          f"{sum(i['qty'] * i['daily_rate'] for c in data['contracts'] for i in c['items']):,.0f}₮")
    print(f"  Нөөц      : {len(data['stock'])} мөр · "
          f"{sum(s['on_hand'] for s in data['stock']):,.0f}ш")
    print(f"  Зээл      : {len(data['loans'])} · "
          f"{sum(l['principal'] for l in data['loans']):,.0f}₮")
    print(f"  Бартер    : {len(data['barter'])}")
    print(f"  Анхааруулга: {len(rep.warnings)} · задлагдаагүй нүд: {len(rep.unparsed)}")
    for w in rep.warnings:
        print("   ⚠", w)
    for u in rep.unparsed:
        print("   ✗", u)
    return 0


if __name__ == "__main__":
    sys.exit(main())
