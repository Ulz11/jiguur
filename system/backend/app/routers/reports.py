"""Тайлан + Excel export/import."""
import io
from datetime import date, timedelta
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import Response
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from ..db import get_db
from .. import models, auth, serializers
from ..services import reports as R
from ..services import billing, loans as L

router = APIRouter(prefix="/api")
guard = auth.require_roles("manager", "finance")
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def report_range(months: int, d_from: str, d_to: str, today: date) -> tuple[date, date]:
    """Тайлангийн муж: огноо өгвөл түүгээр, үгүй бол сүүлийн n сар.

    Буруу огноо, урвуу муж хоёул 400 — чимээгүй 500 болж унахын оронд
    юу нь буруу байгааг хэлнэ.
    """
    if d_from and d_to:
        try:
            f, t = date.fromisoformat(d_from), date.fromisoformat(d_to)
        except ValueError:
            raise HTTPException(400, "Огноо буруу байна (ЖЖЖЖ-СС-ӨӨ)")
        if f > t:
            raise HTTPException(400, "Эхлэх огноо дуусахаасаа хойно байна")
        return f, t
    f = (today.replace(day=1) - timedelta(days=30 * (months - 1))).replace(day=1)
    return f, today


@router.get("/reports")
def reports(months: int = 6, d_from: str = "", d_to: str = "",
            db: Session = Depends(get_db), user=Depends(guard)):
    today = date.today()
    f, t = report_range(months, d_from, d_to, today)
    return {"pnl": R.pnl(db, f, t),
            "months": months,
            "series": R.cashflow_series(db, today, 6),
            "loans_total": L.summary(db, today)["total_debt"]}


def _xlsx(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/reports/export.xlsx")
def export_report(months: int = 6, d_from: str = "", d_to: str = "",
                  db: Session = Depends(get_db), user=Depends(guard)):
    today = date.today()
    f, t = report_range(months, d_from, d_to, today)
    p = R.pnl(db, f, t)
    dt = p["detail"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Ашиг алдагдал"
    ws.append(["Жигүүр Зам ХХК — Ашиг, алдагдлын тайлан"])
    ws.append([f"Хугацаа: {p['from']} — {p['to']}"])
    ws.append([])
    ch = dt["charge"]
    rows = [
        ("ОРЛОГО", ""),
        ("Түрээсийн орлого", p["rent_income"]),
        ("    үүнээс цэвэр түрээс", dt["rent_net"]),
        ("    үүнээс засварын нэхэлт", ch["repair"]),
        ("    үүнээс акталсан бүтээгдэхүүн", ch["writeoff"]),
        ("    үүнээс чөлөөт акт (±)", ch["akt"]),
        *([("    үүнээс задаргаагүй", ch["other"])] if ch["other"] else []),
        ("Худалдааны орлого", p["sale_income"]), ("Механизмын орлого", p["machine_income"]),
        *([(f"    дотоод ажил {p['machine_internal_count']}ш — орлогод ороогүй",
            p["machine_internal"])] if p["machine_internal_count"] else []),
        ("Алдангийн орлого (төлөгдсөн)", p["penalty_income"]),
        ("Нийт орлого", p["total_income"]), ("", ""),
        ("ЗАРДАЛ", ""), ("Механизмын зарлага", p["machine_expense"]),
        ("Цалин", p["salary_expense"]), ("Зээлийн хүү", p["interest_expense"]),
        ("Нийт зардал", p["total_expense"]), ("", ""),
        ("Бартерын хэрэгжсэн үр дүн", p["barter_result"]),
        ("ЦЭВЭР ҮР ДҮН", p["net"]),
    ]
    for r_ in rows:
        ws.append(list(r_))
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    # ---- Задаргаа: тоо бүр яаж гарсан нь мөр мөрөөрөө ----
    wz = wb.create_sheet("Задаргаа")
    wz.column_dimensions["A"].width = 13
    wz.column_dimensions["B"].width = 32
    for col in "CDEFG":
        wz.column_dimensions[col].width = 16

    def section(title: str, header: list, data_rows: list, total=None):
        wz.append([title])
        wz.append(header)
        for r_ in data_rows:
            wz.append(r_)
        if total is not None:
            wz.append(["", "НИЙТ"] + total)
        wz.append([])

    section("ТҮРЭЭСИЙН НЭХЭМЖЛЭЛҮҮД (дууссан циклээр)",
            ["Огноо", "Харилцагч", "Гэрээ", "Нэхэмжлэл", "Түрээс", "Засвар/акт", "Дүн"],
            [[r["date"], r["client"], r["contract_no"], r["no"],
              r["rent"], r["charge"], r["total"]] for r in dt["rent_invoices"]],
            [None, None, dt["rent_net"], sum(ch[k] for k in ("repair", "writeoff", "akt", "other")),
             p["rent_income"]])
    section("ЗАСВАР / АКТЫН МӨРҮҮД",
            ["Огноо", "Харилцагч", "Гэрээ", "Төрөл", "Дүн"],
            [[r["date"], r["client"], r["contract_no"], r["desc"], r["amount"]]
             for r in ch["rows"]])
    section("ХУДАЛДАА",
            ["Огноо", "Харилцагч", "Гэрээ", "Нэхэмжлэл", "Дүн"],
            [[r["date"], r["client"], r["contract_no"], r["no"], r["amount"]]
             for r in dt["sale_invoices"]])
    section("АЛДАНГИ — ТӨЛӨГДСӨН (орлогод орсон)",
            ["Огноо", "Харилцагч", "Нэхэмжлэл", "Дүн"],
            [[r["date"], r["client"], r["invoice_no"], r["amount"]]
             for r in dt["penalty_paid"]])
    section(f"АЛДАНГИ — ЭНЭ ХУГАЦААНД НЭХЭГДСЭН (нийт {dt['penalty_booked']['total']}₮, орлогод ОРООГҮЙ)",
            ["Огноо", "Харилцагч", "Гэрээ", "Дүн", "Хэн нэхсэн"],
            [[r["date"], r["client"], r["contract_no"], r["amount"], r["user"]]
             for r in dt["penalty_booked"]["rows"]])
    section("БАРТЕР — орж ирсэн ↔ зарагдсан",
            ["Хөрөнгө", "Хэнээс", "Орж ирсэн огноо", "Орж ирсэн үнэ",
             "Зарсан огноо", "Хэнд", "Зарсан үнэ", "Зөрүү"],
            [[r["name"], r["client"], r["date_in"], r["value_in"],
              r["sold_date"], r["sold_to"], r["sold_amount"], r["diff"]]
             for r in dt["barter"]])
    section("МЕХАНИЗМ — машин бүрээр",
            ["Машин", "Орлого", "Зарлага", "Цэвэр"],
            [[r["machine"], r["income"], r["expense"], r["net"]]
             for r in dt["machines"]])
    section("ЦАЛИН — олгосон бодолтууд",
            ["Огноо", "Бодолт", "Ажилтан", "Дүн (base)"],
            [[r["date"], r["label"], r["employees"], r["amount"]]
             for r in dt["salary"]])
    section("ЗЭЭЛИЙН ХҮҮ — төлөлт бүрээр",
            ["Огноо", "Зээлдүүлэгч", "Дүн"],
            [[r["date"], r["loan"], r["amount"]] for r in dt["interest"]])

    ws2 = wb.create_sheet("Авлага")
    # Алдангийн ХОЁР багана — нэхэгдсэн нь өр, нэхэгдээгүй нь зөвхөн тооцоолол.
    # Нэг багана болгож нийлүүлбэл Excel рүү буусан тоо нь «нэхсэн» гэж
    # уншигдаж, хэзээ ч гаргаагүй шийдвэр баримт болно (R25 / H2).
    ws2.append(["Харилцагч", "Идэвхтэй гэрээ", "Авлагын үлдэгдэл",
                "Нэхэгдсэн алданги", "Алдангийн тооцоолол (нэхэгдээгүй)", "Барьцаа"])
    today_ = date.today()
    for c in db.query(models.Client).order_by(models.Client.name).all():
        row = serializers.client_row(c, today_)
        ws2.append([row["name"], row["active_contracts"], row["receivable"],
                    row["penalty_booked"], row["penalty_unbooked"], row["deposit"]])
    ws2.column_dimensions["A"].width = 32

    ws3 = wb.create_sheet("Зээл")
    ws3.append(["Зээлдүүлэгч", "Үндсэн дүн", "Нэмэлт олголт", "Үлдэгдэл",
                "Хүү %/сар", "Сарын хүү", "Сарын төлөлт"])
    from ..services.loans import loan_balance, monthly_due, topup_total
    for l in db.query(models.Loan).filter_by(status="active").all():
        ws3.append([l.name, l.principal, topup_total(l), loan_balance(l),
                    l.monthly_rate, monthly_due(l), l.monthly_payment or 0])
    ws3.column_dimensions["A"].width = 32

    return Response(_xlsx(wb), media_type=XLSX_MIME,
                    headers={"Content-Disposition": 'attachment; filename="tailan.xlsx"'})


@router.get("/export/receivables.xlsx")
def export_receivables(db: Session = Depends(get_db), user=Depends(guard)):
    today = date.today()
    for c in db.query(models.Contract).filter_by(status="active").all():
        billing.ensure_invoices(db, c, today)
    wb = Workbook()
    ws = wb.active
    ws.title = "Авлага"
    # Авлагын НИЙТ дүн нь дэлгэц бүрийнхтэй ЯГ ижил (H9b); хажууд нь
    # задаргаа — цаас дээр тулгахад «энэ хоёр тоо яагаад зөрөв» гэж
    # асуух шаардлагагүй байхын тулд.
    ws.append(["Харилцагч", "Регистр", "Утас", "Идэвхтэй гэрээ",
               "Авлагын үлдэгдэл", "үүнээс нэхэмжилсэн",
               "үүнээс нэхэмжлэгдээгүй", "Нэхэгдсэн алданги",
               "Алдангийн тооцоолол (нэхэгдээгүй)", "Барьцаа", "Хэтэрсэн эсэх"])
    for c in db.query(models.Client).order_by(models.Client.name).all():
        row = serializers.client_row(c, today)
        ws.append([row["name"], row["reg"], row["phone"], row["active_contracts"],
                   row["receivable"], row["receivable_invoiced"],
                   row["receivable_uninvoiced"],
                   row["penalty_booked"], row["penalty_unbooked"],
                   row["deposit"], "Тийм" if row["overdue"] else ""])
    ws.column_dimensions["A"].width = 32
    return Response(_xlsx(wb), media_type=XLSX_MIME,
                    headers={"Content-Disposition": 'attachment; filename="avlaga.xlsx"'})


@router.post("/import/clients")
async def import_clients(file: UploadFile, db: Session = Depends(get_db),
                         user=Depends(auth.require_roles("manager", "finance"))):
    """XLSX: Нэр | Регистр | Хариуцагч | Утас (эхний мөр — толгой)."""
    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True)
    except Exception:
        raise HTTPException(400, "XLSX файл уншигдсангүй")
    ws = wb.active
    existing = {c.name.strip().lower() for c in db.query(models.Client).all()}
    created = skipped = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # толгой
        name = str(row[0] or "").strip()
        if not name:
            continue
        if name.lower() in existing:
            skipped += 1
            continue
        db.add(models.Client(name=name,
                             reg=str(row[1] or "").strip() if len(row) > 1 else "",
                             person=str(row[2] or "").strip() if len(row) > 2 else "",
                             phone=str(row[3] or "").strip() if len(row) > 3 else ""))
        existing.add(name.lower())
        created += 1
    db.commit()
    return {"created": created, "skipped": skipped}
